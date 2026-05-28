"""
JNFS Package Inspector
======================

Two main modes:

1. report  — analyze a build manifest (JSON) and produce resource breakdown:
              by extension, by SVT/streaming category, by PackageTag,
              plus default Excel export and Feishu sheet upload.

2. lookup  — given exact Unity AssetPath(s), tell whether each asset is in
              the package, what JNFS file(s) it lives in, its hash and size.

Inputs:
  - manifest_json:  build manifest with AssetPath / Hash / InternalPath /
                    PackageTag / FileSize (e.g. StandaloneWindows64_common_full.json)
  - game_data_dir:  optional. Root that contains the JNFS .idx/.data files
                    (e.g. T3_shiwan/Game_Data/StreamingAssets/res). When provided
                    we cross-check that each manifest Hash lives in a real JNFS
                    DataEntry and resolve which .data file holds it.

Manifest fields used (all from per-entry dict):
  AssetPath, Hash, InternalPath, PackageTag, FileSize

JNFS layout (used only in --game-data mode for cross-check):
  .idx :
    Header  : fileEntryCount (U32 BE)  dataEntryCount (U32 BE)
    DataEntry[] (28 B) : contentHash(16)  offset(U32 BE)  encSize(U32 BE)  decSize(U32 BE)
    FileEntry[] (24 B) : pathHash(U64 BE)  contentHash(16)
"""

from __future__ import annotations

import argparse
import collections
import json
import os
import re
import struct
import sys
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

# ─── Optional Excel ──────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════
# JNFS reader (idx-only — we do NOT decompress .data here, just verify hash
# presence and resolve packages)
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class DataEntry:
    content_hash: bytes
    offset: int
    encoded_size: int
    decoded_size: int


@dataclass
class FileEntry:
    path_hash: int
    content_hash: bytes


def read_idx(idx_path: str) -> Tuple[List[DataEntry], List[FileEntry]]:
    with open(idx_path, "rb") as f:
        raw = f.read()
    if len(raw) < 8:
        raise ValueError(f"idx too short: {idx_path}")
    file_n, data_n = struct.unpack_from(">II", raw, 0)
    off = 8
    data_entries: List[DataEntry] = []
    for _ in range(data_n):
        ch = raw[off:off + 16]
        eo, es, ds = struct.unpack_from(">III", raw, off + 16)
        data_entries.append(DataEntry(ch, eo, es, ds))
        off += 28
    file_entries: List[FileEntry] = []
    for _ in range(file_n):
        ph = struct.unpack_from(">Q", raw, off)[0]
        ch = raw[off + 8: off + 24]
        file_entries.append(FileEntry(ph, ch))
        off += 24
    return data_entries, file_entries


def scan_jnfs_root(game_data_dir: str) -> Dict[str, str]:
    """Return {hex(content_hash) -> package_label(rel path)} index of every
    JNFS file across all .idx files under the root."""
    hash_to_pkg: Dict[str, str] = {}
    idx_files = []
    for root, _, files in os.walk(game_data_dir):
        for fn in files:
            if fn.endswith(".idx"):
                idx_files.append(os.path.join(root, fn))
    idx_files.sort()
    print(f"  Found {len(idx_files)} .idx file(s) under {game_data_dir}", file=sys.stderr)
    for idx in idx_files:
        try:
            data_entries, file_entries = read_idx(idx)
        except Exception as e:
            print(f"  [WARN] {idx}: {e}", file=sys.stderr)
            continue
        # We index by FileEntry's contentHash (= what the manifest 'Hash' joins on)
        # because a single DataEntry can be referenced by many FileEntry rows
        # (deduplication). But here we just want hash -> some package label.
        rel = os.path.relpath(idx, game_data_dir).replace(os.sep, "/")
        for fe in file_entries:
            key = fe.content_hash.hex()
            hash_to_pkg.setdefault(key, rel)
        # Also seed with data_entries in case some file entries are absent.
        for de in data_entries:
            hash_to_pkg.setdefault(de.content_hash.hex(), rel)
    return hash_to_pkg


# ═══════════════════════════════════════════════════════════════════════════
# Manifest loading & normalization
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class ManifestEntry:
    asset_path: str          # may be empty (e.g. SVT files)
    internal_path: str       # always present
    hash: str                # 32-hex MD5
    package_tag: str
    file_size: int

    @property
    def primary_path(self) -> str:
        return self.asset_path or self.internal_path


def load_manifest(path: str) -> List[ManifestEntry]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Manifest root must be a JSON array")
    out: List[ManifestEntry] = []
    for e in data:
        out.append(ManifestEntry(
            asset_path=(e.get("AssetPath") or "").strip(),
            internal_path=(e.get("InternalPath") or "").strip(),
            hash=(e.get("Hash") or "").lower().strip(),
            package_tag=(e.get("PackageTag") or "").strip(),
            file_size=int(e.get("FileSize") or 0),
        ))
    return out


# ═══════════════════════════════════════════════════════════════════════════
# Bucketing — extension + SVT/streaming categories
# ═══════════════════════════════════════════════════════════════════════════

# Special path-prefix categories take precedence over plain extension.
# Order matters: first match wins. Tune the patterns as new package layouts appear.
SPECIAL_BUCKETS: List[Tuple[str, re.Pattern]] = [
    ("svt",            re.compile(r"(^|/)svt/",            re.IGNORECASE)),
    ("streaming",      re.compile(r"(^|/)streaming",       re.IGNORECASE)),
    ("entity-scenes",  re.compile(r"(^|/)entityscenes",    re.IGNORECASE)),
    ("audio",          re.compile(r"(^|/)audio/",          re.IGNORECASE)),
    ("video",          re.compile(r"(^|/)video/",          re.IGNORECASE)),
    ("scene-streaming",re.compile(r"\.scenestreaming",     re.IGNORECASE)),
    ("addressable",    re.compile(r"(^|/)addressable",     re.IGNORECASE)),
]


def bucket_special(entry: ManifestEntry) -> Optional[str]:
    path_to_check = entry.internal_path or entry.asset_path
    if not path_to_check:
        return None
    for name, pat in SPECIAL_BUCKETS:
        if pat.search(path_to_check):
            return name
    return None


def bucket_extension(entry: ManifestEntry) -> str:
    base = (entry.asset_path or entry.internal_path).rsplit("/", 1)[-1]
    if "." not in base:
        # Many SVT entries have no extension because AssetPath is empty;
        # fall back to InternalPath in that case.
        base = entry.internal_path.rsplit("/", 1)[-1]
    if "." in base:
        return "." + base.rsplit(".", 1)[-1].lower()
    return "(no-ext)"


# ═══════════════════════════════════════════════════════════════════════════
# Reports
# ═══════════════════════════════════════════════════════════════════════════

def build_report(entries: List[ManifestEntry],
                 hash_to_pkg: Optional[Dict[str, str]] = None) -> Dict:
    """Return a dict that contains all numbers needed for Markdown + Excel."""
    # Per-extension stats (after special-bucket extraction)
    ext_count = collections.Counter()
    ext_bytes = collections.Counter()
    special_count = collections.Counter()
    special_bytes = collections.Counter()
    pkgtag_count = collections.Counter()
    pkgtag_bytes = collections.Counter()
    # SVT-specific drill-down (per texture set GUID)
    svt_pat = re.compile(r"^svt/([0-9a-f]{32})", re.IGNORECASE)
    svt_set = collections.defaultdict(lambda: {"files": 0, "bytes": 0,
                                               "has_gts": False, "gtp": 0})

    missing_in_jnfs = []   # entries whose hash isn't found in JNFS (only if game_data given)

    for e in entries:
        special = bucket_special(e)
        if special:
            special_count[special] += 1
            special_bytes[special] += e.file_size
            if special == "svt":
                ip = e.internal_path.lower()
                m = svt_pat.match(ip)
                if m:
                    g = m.group(1)
                    svt_set[g]["files"] += 1
                    svt_set[g]["bytes"] += e.file_size
                    if ip.endswith(".gts"):
                        svt_set[g]["has_gts"] = True
                    elif ip.endswith(".gtp"):
                        svt_set[g]["gtp"] += 1
        ext = bucket_extension(e)
        ext_count[ext] += 1
        ext_bytes[ext] += e.file_size

        pkgtag_count[e.package_tag] += 1
        pkgtag_bytes[e.package_tag] += e.file_size

        if hash_to_pkg is not None and e.hash and e.hash not in hash_to_pkg:
            missing_in_jnfs.append(e)

    return {
        "total_entries": len(entries),
        "total_bytes": sum(e.file_size for e in entries),
        "ext_count": dict(ext_count.most_common()),
        "ext_bytes": {k: ext_bytes[k] for k in ext_count},
        "special_count": dict(special_count.most_common()),
        "special_bytes": {k: special_bytes[k] for k in special_count},
        "pkgtag_count": dict(pkgtag_count.most_common()),
        "pkgtag_bytes": {k: pkgtag_bytes[k] for k in pkgtag_count},
        "svt_sets": {g: v for g, v in sorted(svt_set.items(),
                                             key=lambda kv: -kv[1]["bytes"])},
        "missing_in_jnfs_count": len(missing_in_jnfs),
        "missing_in_jnfs_sample": [
            {"asset_path": e.asset_path or e.internal_path,
             "hash": e.hash, "size": e.file_size}
            for e in missing_in_jnfs[:30]
        ],
    }


def _mb(n: int) -> str:
    return f"{n / 1024 / 1024:,.2f} MB"


def _gb(n: int) -> str:
    return f"{n / 1024 / 1024 / 1024:,.2f} GB"


def write_markdown(report: Dict, out_path: str, source_label: str) -> None:
    lines: List[str] = []
    lines.append(f"# JNFS Package Inspect — {source_label}\n")
    lines.append(f"- Total entries: **{report['total_entries']:,}**")
    lines.append(f"- Total bytes:   **{_gb(report['total_bytes'])}** "
                 f"({report['total_bytes']:,} bytes)\n")

    if report["missing_in_jnfs_count"]:
        lines.append(f"> ⚠ {report['missing_in_jnfs_count']:,} manifest entries "
                     "have a Hash that does NOT appear in the scanned JNFS .idx files. "
                     "See `missing_in_jnfs` section below.\n")

    # PackageTag
    lines.append("## By PackageTag\n")
    lines.append("| PackageTag | Files | Total Size |")
    lines.append("|---|---:|---:|")
    for tag, n in report["pkgtag_count"].items():
        lines.append(f"| `{tag or '(empty)'}` | {n:,} | {_mb(report['pkgtag_bytes'][tag])} |")
    lines.append("")

    # Special buckets
    if report["special_count"]:
        lines.append("## Special buckets (SVT / Streaming / Audio / …)\n")
        lines.append("| Bucket | Files | Total Size |")
        lines.append("|---|---:|---:|")
        for k, n in report["special_count"].items():
            lines.append(f"| `{k}` | {n:,} | {_mb(report['special_bytes'][k])} |")
        lines.append("")

    # SVT drill-down
    if report["svt_sets"]:
        lines.append("### SVT texture sets (per GUID)\n")
        lines.append("| Set GUID | Files | .gts | .gtp | Total Size |")
        lines.append("|---|---:|:-:|---:|---:|")
        for g, v in report["svt_sets"].items():
            lines.append(f"| `{g}` | {v['files']} | "
                         f"{'✔' if v['has_gts'] else '–'} | "
                         f"{v['gtp']} | {_mb(v['bytes'])} |")
        lines.append("")

    # Extension table
    lines.append("## By extension\n")
    lines.append("| Ext | Files | Total Size |")
    lines.append("|---|---:|---:|")
    for ext, n in report["ext_count"].items():
        lines.append(f"| `{ext}` | {n:,} | {_mb(report['ext_bytes'][ext])} |")
    lines.append("")

    # Missing in JNFS (only if cross-checked)
    if report["missing_in_jnfs_count"]:
        lines.append("## Missing in JNFS (sample)\n")
        lines.append("Manifest claims these assets are in the build, but the Hash "
                     "is not present in any scanned `.idx` file:\n")
        lines.append("| AssetPath | Hash | Size |")
        lines.append("|---|---|---:|")
        for m in report["missing_in_jnfs_sample"]:
            lines.append(f"| `{m['asset_path']}` | `{m['hash']}` | {m['size']:,} |")
        lines.append("")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def write_xlsx(report: Dict, entries: List[ManifestEntry],
               out_path: str, source_label: str,
               hash_to_pkg: Optional[Dict[str, str]] = None) -> None:
    if not HAS_OPENPYXL:
        print("  [SKIP xlsx] openpyxl not installed; run: pip install openpyxl",
              file=sys.stderr)
        return
    wb = Workbook()
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")

    def _new_sheet(title: str, headers: List[str]):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
            c.fill = header_fill
        return ws

    # Drop default sheet
    wb.remove(wb.active)

    # Overview
    ws = _new_sheet("Overview", ["Metric", "Value"])
    ws.append(["Source", source_label])
    ws.append(["Total entries", report["total_entries"]])
    ws.append(["Total bytes", report["total_bytes"]])
    ws.append(["Total MB", round(report["total_bytes"] / 1024 / 1024, 2)])
    ws.append(["Total GB", round(report["total_bytes"] / 1024 / 1024 / 1024, 2)])
    if hash_to_pkg is not None:
        ws.append(["Manifest entries missing in JNFS", report["missing_in_jnfs_count"]])

    # By PackageTag
    ws = _new_sheet("ByPackageTag", ["PackageTag", "Files", "Bytes", "MB"])
    for tag, n in report["pkgtag_count"].items():
        b = report["pkgtag_bytes"][tag]
        ws.append([tag, n, b, round(b / 1024 / 1024, 2)])

    # Special
    ws = _new_sheet("SpecialBuckets", ["Bucket", "Files", "Bytes", "MB"])
    for k, n in report["special_count"].items():
        b = report["special_bytes"][k]
        ws.append([k, n, b, round(b / 1024 / 1024, 2)])

    # SVT sets
    ws = _new_sheet("SVTSets", ["SetGUID", "Files", "HasGTS", "GTP", "Bytes", "MB"])
    for g, v in report["svt_sets"].items():
        ws.append([g, v["files"], "Y" if v["has_gts"] else "N",
                   v["gtp"], v["bytes"], round(v["bytes"] / 1024 / 1024, 2)])

    # Extensions
    ws = _new_sheet("ByExtension", ["Ext", "Files", "Bytes", "MB"])
    for ext, n in report["ext_count"].items():
        b = report["ext_bytes"][ext]
        ws.append([ext, n, b, round(b / 1024 / 1024, 2)])

    # Full entry list (capped — Excel max 1,048,576 rows but huge files are slow)
    ws = _new_sheet("AllEntries",
                    ["AssetPath", "InternalPath", "PackageTag",
                     "Extension", "SpecialBucket", "FileSize", "Hash",
                     "InJNFSPackage"])
    for e in entries:
        ws.append([
            e.asset_path,
            e.internal_path,
            e.package_tag,
            bucket_extension(e),
            bucket_special(e) or "",
            e.file_size,
            e.hash,
            (hash_to_pkg or {}).get(e.hash, "") if hash_to_pkg is not None else "",
        ])

    # Auto-ish column widths
    for ws in wb.worksheets:
        for col_idx, col in enumerate(ws.columns, 1):
            try:
                values = [str(c.value) if c.value is not None else "" for c in col]
            except Exception:
                continue
            width = min(max((len(v) for v in values[:200]), default=10) + 2, 80)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════
# Lookup — is asset X in the package?
# ═══════════════════════════════════════════════════════════════════════════

def lookup_assets(entries: List[ManifestEntry], queries: List[str],
                  hash_to_pkg: Optional[Dict[str, str]] = None) -> List[Dict]:
    """Exact AssetPath match (case-sensitive). Returns one result per query."""
    by_path = {e.asset_path: e for e in entries if e.asset_path}
    results = []
    for q in queries:
        q = q.strip()
        e = by_path.get(q)
        if e is None:
            results.append({"query": q, "in_package": False,
                            "reason": "AssetPath not in manifest"})
            continue
        rec = {
            "query": q,
            "in_package": True,
            "asset_path": e.asset_path,
            "internal_path": e.internal_path,
            "package_tag": e.package_tag,
            "file_size": e.file_size,
            "hash": e.hash,
            "extension": bucket_extension(e),
            "special_bucket": bucket_special(e) or "",
        }
        if hash_to_pkg is not None:
            pkg = hash_to_pkg.get(e.hash)
            rec["jnfs_package"] = pkg or "(not found in JNFS)"
            rec["jnfs_verified"] = pkg is not None
        results.append(rec)
    return results


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def cmd_report(args: argparse.Namespace) -> int:
    print(f"[load] {args.manifest}", file=sys.stderr)
    entries = load_manifest(args.manifest)
    print(f"[load] {len(entries):,} entries", file=sys.stderr)

    hash_to_pkg = None
    if args.game_data:
        print(f"[scan-jnfs] {args.game_data}", file=sys.stderr)
        hash_to_pkg = scan_jnfs_root(args.game_data)
        print(f"[scan-jnfs] {len(hash_to_pkg):,} distinct hashes", file=sys.stderr)

    report = build_report(entries, hash_to_pkg)

    os.makedirs(args.output, exist_ok=True)
    label = os.path.basename(args.manifest)

    md = os.path.join(args.output, "report.md")
    write_markdown(report, md, label)
    print(f"[write] {md}", file=sys.stderr)

    raw = os.path.join(args.output, "report.json")
    with open(raw, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"[write] {raw}", file=sys.stderr)

    if not args.no_xlsx:
        xlsx = os.path.join(args.output, "report.xlsx")
        write_xlsx(report, entries, xlsx, label, hash_to_pkg)
        if os.path.isfile(xlsx):
            print(f"[write] {xlsx}", file=sys.stderr)

    # Console summary
    print()
    print(f"=== Summary: {label} ===")
    print(f"Total entries : {report['total_entries']:>10,}")
    print(f"Total size    : {_gb(report['total_bytes']):>10}")
    print(f"Special buckets:")
    for k, n in report["special_count"].items():
        print(f"  {k:<16} {n:>8,}  {_mb(report['special_bytes'][k])}")
    print(f"Top 10 extensions:")
    for ext, n in list(report["ext_count"].items())[:10]:
        print(f"  {ext:<16} {n:>8,}  {_mb(report['ext_bytes'][ext])}")
    if hash_to_pkg is not None and report["missing_in_jnfs_count"]:
        print(f"⚠ {report['missing_in_jnfs_count']:,} manifest hashes not found in JNFS")
    return 0


def cmd_lookup(args: argparse.Namespace) -> int:
    entries = load_manifest(args.manifest)
    queries = list(args.asset_path) if args.asset_path else []
    if args.from_file:
        with open(args.from_file, "r", encoding="utf-8") as f:
            for ln in f:
                ln = ln.strip()
                if ln and not ln.startswith("#"):
                    queries.append(ln)
    if not queries:
        print("ERROR: provide --asset-path or --from-file", file=sys.stderr)
        return 2

    hash_to_pkg = None
    if args.game_data:
        hash_to_pkg = scan_jnfs_root(args.game_data)

    results = lookup_assets(entries, queries, hash_to_pkg)

    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        for r in results:
            if r["in_package"]:
                line = (f"[IN ] {r['query']}\n"
                        f"      hash={r['hash']}  size={r['file_size']:,}  "
                        f"pkg_tag={r['package_tag']}  ext={r['extension']}  "
                        f"special={r['special_bucket'] or '-'}")
                if "jnfs_package" in r:
                    line += f"\n      jnfs={r['jnfs_package']}"
                print(line)
            else:
                print(f"[OUT] {r['query']}   ({r['reason']})")
    # Exit non-zero if any query missed — handy for CI/scripted checks
    return 0 if all(r["in_package"] for r in results) else 1


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="JNFS package inspector")
    sp = p.add_subparsers(dest="cmd", required=True)

    pr = sp.add_parser("report", help="Generate resource-type breakdown report")
    pr.add_argument("manifest", help="Path to build manifest JSON")
    pr.add_argument("-o", "--output", default="JnfsReport",
                    help="Output directory (default: JnfsReport)")
    pr.add_argument("--game-data", default=None,
                    help="Optional: Game_Data root containing JNFS .idx files "
                         "(cross-checks manifest Hash against actual packages)")
    pr.add_argument("--no-xlsx", action="store_true",
                    help="Skip Excel export")
    pr.set_defaults(func=cmd_report)

    pl = sp.add_parser("lookup", help="Check whether assets are in the package")
    pl.add_argument("manifest", help="Path to build manifest JSON")
    pl.add_argument("--asset-path", action="append", default=[],
                    help="Exact AssetPath to look up (repeatable)")
    pl.add_argument("--from-file", default=None,
                    help="Text file with one AssetPath per line")
    pl.add_argument("--game-data", default=None,
                    help="Optional: cross-check each hit against real JNFS data")
    pl.add_argument("--json", action="store_true",
                    help="Emit machine-readable JSON")
    pl.set_defaults(func=cmd_lookup)

    args = p.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
