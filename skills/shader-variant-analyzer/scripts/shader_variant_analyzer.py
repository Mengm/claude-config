#!/usr/bin/env python3
"""
Shader Variant Analyzer — JNFS Package Shader Parser (Route B)

解析已打包的 JNFS 包体 + 标准 UnityFS (data.unity3d)，提取 Shader 对象的变体信息。
支持 JNTE XOR 混淆的 zstd 解压。

用法:
    python shader_variant_analyzer.py <game_data_dir> [-o output_dir] [--project-dir <unity_project_dir>]

示例:
    python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" -o ShaderReport
    python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" \
        --project-dir "Project-T3-zhengyang.solis-development/client" -o ShaderReport

依赖:
    pip install zstandard lz4 openpyxl
"""

import argparse
import hashlib
import io
import json
import os
import struct
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Dict, List, Optional, Tuple

try:
    import zstandard as zstd
except ImportError:
    zstd = None

try:
    import lz4.block
except ImportError:
    lz4 = None

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ════════════════════════════════════════════════════════════════════
# Constants
# ════════════════════════════════════════════════════════════════════

# XOR masks from jnunity/External/Compression/zstd/common/mem.h
# In GCC/Clang, multi-char constants: 'JN' = (J<<8)|N, 'JNTE' = (J<<24)|(N<<16)|(T<<8)|E
XOR_U16 = 0x4A4E          # (U16)'JN'  = (0x4A << 8) | 0x4E
XOR_U32 = 0x4A4E5445      # (U32)'JNTE' = (0x4A << 24) | (0x4E << 16) | (0x54 << 8) | 0x45
XOR_U64 = 0x637c777bf26b6fc5

# Standard zstd magic (upstream)
ZSTD_MAGIC_STANDARD = 0xFD2FB528
# JN engine's custom zstd magic (modified in zstd.h, NOT XOR'd — directly replaced)
ZSTD_MAGIC_JN = 0xFA4BF25C

# JNTE constants
JNTE_SIGNATURE = 0x4A4E5445  # 'JNTE' as big-endian U32
CHUNK_RAW = 0
CHUNK_LZ4 = 1
CHUNK_LZMA = 2
CHUNK_ZSTD = 3

# Unity constants
UNITY_SHADER_CLASS_ID = 48

# UnityFS magic
UNITYFS_MAGIC = b"UnityFS"


# ════════════════════════════════════════════════════════════════════
# Data classes
# ════════════════════════════════════════════════════════════════════

@dataclass
class JNFSDataEntry:
    content_hash: bytes  # 16 bytes
    offset: int          # U32
    encoded_size: int    # U32
    decoded_size: int    # U32


@dataclass
class JNFSFileEntry:
    path_hash: int       # U64
    content_hash: bytes  # 16 bytes


@dataclass
class SubProgramInfo:
    gpu_program_type: int
    keyword_indices: List[int]
    blob_index: int
    source_kind: str
    player_tier: Optional[int] = None
    graphics_tier: Optional[int] = None


@dataclass
class PassInfo:
    name: str
    programs: Dict[str, List[SubProgramInfo]] = field(default_factory=dict)
    # programs key: stage name (Vertex, Fragment, etc.)


@dataclass
class SubShaderInfo:
    lod: int
    passes: List[PassInfo] = field(default_factory=list)


@dataclass
class ShaderInfo:
    name: str
    keywords: List[str] = field(default_factory=list)
    keyword_flags: List[int] = field(default_factory=list)
    sub_shaders: List[SubShaderInfo] = field(default_factory=list)
    source_file: str = ""

    @property
    def total_variants(self) -> int:
        count = 0
        for ss in self.sub_shaders:
            for p in ss.passes:
                for stage, subs in p.programs.items():
                    count += len(subs)
        return count

    @property
    def total_unique_variants(self) -> int:
        count = 0
        for ss in self.sub_shaders:
            for p in ss.passes:
                for stage, subs in p.programs.items():
                    count += count_unique_subprograms(subs)
        return count


def subprogram_identity(sp: SubProgramInfo) -> tuple:
    return (
        sp.gpu_program_type,
        tuple(sp.keyword_indices),
        sp.blob_index,
        sp.source_kind,
        sp.player_tier,
        sp.graphics_tier,
    )


def count_unique_subprograms(subs: List[SubProgramInfo]) -> int:
    return len({subprogram_identity(sp) for sp in subs})


# ════════════════════════════════════════════════════════════════════
# JNTE Zstd XOR Deobfuscation
# ════════════════════════════════════════════════════════════════════

def fix_zstd_xor(data: bytes) -> bytes:
    """
    Reverse the JN engine's zstd modifications to produce standard zstd data.

    The engine made TWO changes to the zstd format:
    1. Changed the magic number from 0xFD2FB528 to 0xFA4BF25C (in zstd.h)
    2. XOR'd specific frame header fields (FCS, dictID) using _jn read/write functions

    Block data is NOT affected.
    """
    if len(data) < 4:
        return data

    buf = bytearray(data)

    # Check magic: the JN engine uses custom magic 0xFA4BF25C (not XOR'd, directly replaced)
    magic = struct.unpack_from('<I', buf, 0)[0]
    if magic == ZSTD_MAGIC_STANDARD:
        return data  # Already standard zstd
    if magic != ZSTD_MAGIC_JN:
        return data  # Unknown format, return as-is

    # Replace JN magic with standard magic
    struct.pack_into('<I', buf, 0, ZSTD_MAGIC_STANDARD)

    if len(buf) < 5:
        return bytes(buf)

    # Parse frame header descriptor (byte 4)
    fhd_byte = buf[4]
    dict_id_size_code = fhd_byte & 0x03
    single_segment = (fhd_byte >> 5) & 1
    fcs_id = fhd_byte >> 6

    pos = 5  # after magic(4) + fhd(1)

    # Window descriptor (1 byte, not XOR'd) — skip if not single segment
    if not single_segment:
        pos += 1

    # Dictionary ID field
    dict_id_sizes = {0: 0, 1: 1, 2: 2, 3: 4}
    dict_id_size = dict_id_sizes[dict_id_size_code]
    if dict_id_size == 2 and pos + 2 <= len(buf):
        val = struct.unpack_from('<H', buf, pos)[0] ^ XOR_U16
        struct.pack_into('<H', buf, pos, val)
    elif dict_id_size == 4 and pos + 4 <= len(buf):
        val = struct.unpack_from('<I', buf, pos)[0] ^ XOR_U32
        struct.pack_into('<I', buf, pos, val)
    # dict_id_size 0 or 1: no XOR applied (1-byte has no _jn variant)
    pos += dict_id_size

    # Frame Content Size field
    if fcs_id == 0:
        # 0 or 1 byte (1 byte if single_segment) — no XOR for single byte
        if single_segment:
            pos += 1
    elif fcs_id == 1 and pos + 2 <= len(buf):
        val = struct.unpack_from('<H', buf, pos)[0] ^ XOR_U16
        struct.pack_into('<H', buf, pos, val)
        pos += 2
    elif fcs_id == 2 and pos + 4 <= len(buf):
        val = struct.unpack_from('<I', buf, pos)[0] ^ XOR_U32
        struct.pack_into('<I', buf, pos, val)
        pos += 4
    elif fcs_id == 3 and pos + 8 <= len(buf):
        val = struct.unpack_from('<Q', buf, pos)[0] ^ XOR_U64
        struct.pack_into('<Q', buf, pos, val)
        pos += 8

    return bytes(buf)


# ════════════════════════════════════════════════════════════════════
# JNTE Decompressor
# ════════════════════════════════════════════════════════════════════

class JNTEDecompressor:
    """Decompress JNTE containers with support for XOR'd zstd."""

    @staticmethod
    def decompress(raw_blob: bytes) -> bytes:
        """Decompress a JNTE blob into raw data."""
        if len(raw_blob) < 8:
            raise ValueError(f"JNTE blob too short: {len(raw_blob)} bytes")

        # Read JNTE header (big-endian)
        sig, chunk_count = struct.unpack_from('>II', raw_blob, 0)
        if sig != JNTE_SIGNATURE:
            raise ValueError(f"Invalid JNTE signature: 0x{sig:08X} (expected 0x{JNTE_SIGNATURE:08X})")

        # Read chunk info entries (9 bytes each: type(1) + decompressedSize(4) + compressedSize(4), big-endian)
        offset = 8
        chunks = []
        for i in range(chunk_count):
            if offset + 9 > len(raw_blob):
                raise ValueError(f"JNTE truncated at chunk info {i}")
            ctype = raw_blob[offset]
            decomp_size, comp_size = struct.unpack_from('>II', raw_blob, offset + 1)
            chunks.append((ctype, decomp_size, comp_size))
            offset += 9

        # Decompress each chunk
        result = bytearray()
        for i, (ctype, decomp_size, comp_size) in enumerate(chunks):
            if offset + comp_size > len(raw_blob):
                raise ValueError(f"JNTE truncated at chunk data {i}: need {comp_size} bytes at offset {offset}, have {len(raw_blob)}")
            chunk_data = raw_blob[offset:offset + comp_size]
            offset += comp_size

            if ctype == CHUNK_RAW:
                result.extend(chunk_data)
            elif ctype == CHUNK_LZ4:
                if lz4 is None:
                    raise ImportError("lz4 package required: pip install lz4")
                decompressed = lz4.block.decompress(chunk_data, uncompressed_size=decomp_size)
                result.extend(decompressed)
            elif ctype == CHUNK_LZMA:
                import lzma
                decompressed = lzma.decompress(chunk_data)
                result.extend(decompressed)
            elif ctype == CHUNK_ZSTD:
                if zstd is None:
                    raise ImportError("zstandard package required: pip install zstandard")
                # Fix XOR obfuscation in zstd frame header
                fixed_data = fix_zstd_xor(chunk_data)
                dctx = zstd.ZstdDecompressor()
                decompressed = dctx.decompress(fixed_data, max_output_size=decomp_size)
                result.extend(decompressed)
            else:
                raise ValueError(f"Unknown JNTE chunk type: {ctype}")

        return bytes(result)


# ════════════════════════════════════════════════════════════════════
# JNFS Reader
# ════════════════════════════════════════════════════════════════════

class JNFSReader:
    """Parse JNFS .idx + .data archive files."""

    @staticmethod
    def read_idx(idx_path: str) -> Tuple[List[JNFSDataEntry], List[JNFSFileEntry]]:
        """Parse a .idx file, returning (data_entries, file_entries)."""
        with open(idx_path, 'rb') as f:
            data = f.read()

        if len(data) < 8:
            raise ValueError(f"IDX file too short: {len(data)}")

        # Header: fileEntryCount(U32 BE) + dataEntryCount(U32 BE)
        file_entry_count, data_entry_count = struct.unpack_from('>II', data, 0)
        offset = 8

        # DataEntry[]: 28 bytes each (contentHash(16) + offset(U32 BE) + encodedSize(U32 BE) + decodedSize(U32 BE))
        data_entries = []
        for i in range(data_entry_count):
            if offset + 28 > len(data):
                raise ValueError(f"IDX truncated at data entry {i}")
            content_hash = data[offset:offset + 16]
            eo, es, ds = struct.unpack_from('>III', data, offset + 16)
            data_entries.append(JNFSDataEntry(content_hash, eo, es, ds))
            offset += 28

        # FileEntry[]: 24 bytes each (pathHash(U64 BE) + contentHash(16))
        file_entries = []
        for i in range(file_entry_count):
            if offset + 24 > len(data):
                raise ValueError(f"IDX truncated at file entry {i}")
            path_hash = struct.unpack_from('>Q', data, offset)[0]
            content_hash = data[offset + 8:offset + 24]
            file_entries.append(JNFSFileEntry(path_hash, content_hash))
            offset += 24

        return data_entries, file_entries

    @staticmethod
    def read_file(data_path: str, entry: JNFSDataEntry) -> bytes:
        """Read and decompress a single file from a .data archive."""
        with open(data_path, 'rb') as f:
            f.seek(entry.offset)
            raw = f.read(entry.encoded_size)

        if len(raw) < 8:
            return raw

        # Check for JNTE signature
        sig = struct.unpack_from('>I', raw, 0)[0]
        if sig == JNTE_SIGNATURE:
            return JNTEDecompressor.decompress(raw)
        else:
            # Raw data (no JNTE container)
            return raw

    @staticmethod
    def iter_files(idx_path: str, data_path: str):
        """Iterate over all files in a JNFS package, yielding (data_entry, decompressed_bytes)."""
        data_entries, file_entries = JNFSReader.read_idx(idx_path)

        # Build hash → data_entry lookup
        hash_to_entry = {e.content_hash: e for e in data_entries}

        for fe in file_entries:
            de = hash_to_entry.get(fe.content_hash)
            if de is None:
                continue
            try:
                data = JNFSReader.read_file(data_path, de)
                yield de, data
            except Exception as e:
                print(f"  [WARN] Failed to read entry at offset {de.offset}: {e}", file=sys.stderr)


# ════════════════════════════════════════════════════════════════════
# UnityFS Reader
# ════════════════════════════════════════════════════════════════════

class UnityFSReader:
    """Parse standard UnityFS archives (data.unity3d)."""

    @staticmethod
    def read_archive(path: str) -> List[bytes]:
        """Read a UnityFS archive and return list of serialized file data blocks."""
        with open(path, 'rb') as f:
            # Magic
            magic = b""
            while True:
                c = f.read(1)
                if c == b'\x00' or c == b'':
                    break
                magic += c

            if magic != UNITYFS_MAGIC:
                raise ValueError(f"Not a UnityFS file: magic={magic}")

            # Version
            version = struct.unpack('>I', f.read(4))[0]
            # Unity version string
            unity_version = b""
            while True:
                c = f.read(1)
                if c == b'\x00' or c == b'':
                    break
                unity_version += c
            # Generator version string
            generator_version = b""
            while True:
                c = f.read(1)
                if c == b'\x00' or c == b'':
                    break
                generator_version += c

            # Archive header
            file_size = struct.unpack('>Q', f.read(8))[0]
            compressed_block_size = struct.unpack('>I', f.read(4))[0]
            uncompressed_block_size = struct.unpack('>I', f.read(4))[0]
            flags = struct.unpack('>I', f.read(4))[0]

            compression_type = flags & 0x3F
            has_directory_info = (flags & 0x40) != 0
            blocks_at_end = (flags & 0x80) != 0
            padding_at_start = (flags & 0x200) != 0

            # Handle padding alignment (kBlockInfoNeedPaddingAtStart)
            if padding_at_start:
                current = f.tell()
                aligned = (current + 15) & ~15  # align to 16
                if aligned > current:
                    f.seek(aligned)

            # Read block info data
            if blocks_at_end:
                current_pos = f.tell()
                f.seek(file_size - compressed_block_size)
                block_info_data = f.read(compressed_block_size)
                f.seek(current_pos)
            else:
                block_info_data = f.read(compressed_block_size)

            data_start = f.tell()
            # Align data start to 16 bytes if padding flag set
            if padding_at_start:
                aligned = (data_start + 15) & ~15
                if aligned > data_start:
                    data_start = aligned

            # Decompress block info
            if compression_type == 0:
                block_info = block_info_data
            elif compression_type == 1:
                import lzma
                block_info = lzma.decompress(block_info_data)
            elif compression_type in (2, 3):
                if lz4 is None:
                    raise ImportError("lz4 package required: pip install lz4")
                block_info = lz4.block.decompress(block_info_data, uncompressed_size=uncompressed_block_size)
            else:
                raise ValueError(f"Unknown UnityFS compression type: {compression_type}")

            # Parse block info
            bio = io.BytesIO(block_info)
            # Uncompressed data hash (16 bytes)
            bio.read(16)

            # Block storage info count
            block_count = struct.unpack('>I', bio.read(4))[0]
            blocks = []
            for _ in range(block_count):
                uncomp_size = struct.unpack('>I', bio.read(4))[0]
                comp_size = struct.unpack('>I', bio.read(4))[0]
                block_flags = struct.unpack('>H', bio.read(2))[0]
                blocks.append((uncomp_size, comp_size, block_flags))

            # Directory info count
            node_count = struct.unpack('>I', bio.read(4))[0]
            nodes = []
            for _ in range(node_count):
                node_offset = struct.unpack('>Q', bio.read(8))[0]
                node_size = struct.unpack('>Q', bio.read(8))[0]
                node_flags = struct.unpack('>I', bio.read(4))[0]
                # Node name (null-terminated string)
                name = b""
                while True:
                    c = bio.read(1)
                    if c == b'\x00' or c == b'':
                        break
                    name += c
                nodes.append((node_offset, node_size, node_flags, name.decode('utf-8', errors='replace')))

            # Read and decompress all data blocks
            f.seek(data_start)
            decompressed_data = bytearray()
            for uncomp_size, comp_size, block_flags in blocks:
                block_data = f.read(comp_size)
                block_comp = block_flags & 0x3F
                if block_comp == 0:
                    decompressed_data.extend(block_data)
                elif block_comp in (2, 3):
                    if lz4 is None:
                        raise ImportError("lz4 package required")
                    decompressed_data.extend(lz4.block.decompress(block_data, uncompressed_size=uncomp_size))
                elif block_comp == 1:
                    import lzma
                    decompressed_data.extend(lzma.decompress(block_data))
                else:
                    print(f"  [WARN] Unknown block compression {block_comp}, skipping", file=sys.stderr)
                    decompressed_data.extend(b'\x00' * uncomp_size)

            decompressed_data = bytes(decompressed_data)

            # Extract node data
            result = []
            for node_offset, node_size, node_flags, name in nodes:
                result.append(decompressed_data[node_offset:node_offset + node_size])

            return result


# ════════════════════════════════════════════════════════════════════
# Unity SerializedFile Parser
# ════════════════════════════════════════════════════════════════════

class BinaryReader:
    """Helper for reading binary data with position tracking."""

    def __init__(self, data: bytes, big_endian: bool = False):
        self.data = data
        self.pos = 0
        self.big_endian = big_endian
        self._endian = '>' if big_endian else '<'

    def read(self, size: int) -> bytes:
        if self.pos + size > len(self.data):
            raise EOFError(f"read({size}) at {self.pos}, buf size {len(self.data)}")
        result = self.data[self.pos:self.pos + size]
        self.pos += size
        return result

    def read_u8(self) -> int:
        val = self.data[self.pos]
        self.pos += 1
        return val

    def read_u16(self) -> int:
        val = struct.unpack_from(f'{self._endian}H', self.data, self.pos)[0]
        self.pos += 2
        return val

    def read_i16(self) -> int:
        val = struct.unpack_from(f'{self._endian}h', self.data, self.pos)[0]
        self.pos += 2
        return val

    def read_u32(self) -> int:
        if self.pos + 4 > len(self.data):
            raise EOFError(f"read_u32 at {self.pos}, buf size {len(self.data)}")
        val = struct.unpack_from(f'{self._endian}I', self.data, self.pos)[0]
        self.pos += 4
        return val

    def read_i32(self) -> int:
        if self.pos + 4 > len(self.data):
            raise EOFError(f"read_i32 at {self.pos}, buf size {len(self.data)}")
        val = struct.unpack_from(f'{self._endian}i', self.data, self.pos)[0]
        self.pos += 4
        return val

    def read_u64(self) -> int:
        if self.pos + 8 > len(self.data):
            raise EOFError(f"read_u64 at {self.pos}, buf size {len(self.data)}")
        val = struct.unpack_from(f'{self._endian}Q', self.data, self.pos)[0]
        self.pos += 8
        return val

    def read_i64(self) -> int:
        if self.pos + 8 > len(self.data):
            raise EOFError(f"read_i64 at {self.pos}, buf size {len(self.data)}")
        val = struct.unpack_from(f'{self._endian}q', self.data, self.pos)[0]
        self.pos += 8
        return val

    def read_f32(self) -> float:
        val = struct.unpack_from(f'{self._endian}f', self.data, self.pos)[0]
        self.pos += 4
        return val

    def read_string(self) -> str:
        """Read a length-prefixed string (4-byte length + chars), then align to 4."""
        length = self.read_u32()
        s = self.data[self.pos:self.pos + length].decode('utf-8', errors='replace')
        self.pos += length
        self.align(4)
        return s

    def read_cstring(self) -> str:
        """Read a null-terminated string."""
        end = self.data.index(0, self.pos)
        s = self.data[self.pos:end].decode('utf-8', errors='replace')
        self.pos = end + 1
        return s

    def align(self, alignment: int):
        remainder = self.pos % alignment
        if remainder:
            self.pos += alignment - remainder

    def remaining(self) -> int:
        return len(self.data) - self.pos

    def seek(self, pos: int):
        self.pos = pos

    def tell(self) -> int:
        return self.pos


# ════════════════════════════════════════════════════════════════════
# Structured Shader Reader — SerializedShader binary parser
# ════════════════════════════════════════════════════════════════════

# Stage name constants
SHADER_STAGE_NAMES = ['Vertex', 'Fragment', 'Geometry', 'Hull', 'Domain', 'RayTracing']


class StructuredShaderReader:
    """结构化解析 Unity SerializedShader 二进制数据。

    按照 SerializedShaderData.h 的 Transfer() 顺序逐字段读取，
    统计每个 shader 的变体数（SerializedPlayerSubProgram）。
    """

    def read_shader(self, data: bytes, is_big_endian: bool) -> ShaderInfo:
        """从 Shader 对象数据解析出完整的 ShaderInfo。"""
        r = BinaryReader(data, big_endian=is_big_endian)

        # ── Outer NamedObject.m_Name (player build, no m_ObjectHideFlags) ──
        outer_name = r.read_string()

        # ── m_ParsedForm (SerializedShader) ──
        # m_PropInfo (SerializedProperties: just m_Props array)
        self._read_prop_info(r)

        # m_SubShaders[]
        sub_shaders_data = self._read_sub_shaders(r)

        # m_KeywordNames[]
        keyword_names = self._read_string_array(r)

        # m_KeywordFlags[] — may not exist (DidReadLastProperty detection)
        keyword_flags = self._try_read_keyword_flags(r, len(keyword_names))

        # m_Name — the real shader name
        shader_name = r.read_string()

        # We have enough info, skip remaining fields

        # ── Build ShaderInfo ──
        info = ShaderInfo(name=shader_name or outer_name or "<unknown>")
        info.keywords = keyword_names
        info.keyword_flags = keyword_flags

        for ss_data in sub_shaders_data:
            ss_info = SubShaderInfo(lod=ss_data['lod'])
            for pass_data in ss_data['passes']:
                pass_info = PassInfo(name=pass_data.get('name', ''))
                for stage, subs in pass_data.get('programs', {}).items():
                    pass_info.programs[stage] = subs
                ss_info.passes.append(pass_info)
            info.sub_shaders.append(ss_info)

        return info

    # ── String / Array helpers ──

    def _read_string_array(self, r: BinaryReader) -> List[str]:
        """Read Array<String>: S32 count + [String...]"""
        count = r.read_i32()
        if count < 0 or count > 100000:
            raise ValueError(f"Implausible string array count: {count}")
        result = []
        for _ in range(count):
            result.append(r.read_string())
        return result

    def _read_u16_array(self, r: BinaryReader) -> List[int]:
        """Read Array<UInt16>: S32 count + [U16...] + align4"""
        count = r.read_i32()
        if count < 0 or count > 1000000:
            raise ValueError(f"Implausible u16 array count: {count}")
        result = []
        for _ in range(count):
            result.append(r.read_u16())
        r.align(4)
        return result

    def _read_u32_array(self, r: BinaryReader) -> List[int]:
        """Read Array<UInt32>: S32 count + [U32...]"""
        count = r.read_i32()
        if count < 0 or count > 1000000:
            raise ValueError(f"Implausible u32 array count: {count}")
        result = []
        for _ in range(count):
            result.append(r.read_u32())
        return result

    def _read_u8_array(self, r: BinaryReader) -> List[int]:
        """Read Array<UInt8>: S32 count + [U8...] + align4"""
        count = r.read_i32()
        if count < 0 or count > 1000000:
            raise ValueError(f"Implausible u8 array count: {count}")
        result = []
        for _ in range(count):
            result.append(r.read_u8())
        r.align(4)
        return result

    def _read_map_str_str(self, r: BinaryReader) -> Dict[str, str]:
        """Read Map<String, String>: S32 count + [{String, String}...]"""
        count = r.read_i32()
        if count < 0 or count > 100000:
            raise ValueError(f"Implausible map count: {count}")
        result = {}
        for _ in range(count):
            k = r.read_string()
            v = r.read_string()
            result[k] = v
        return result

    def _read_map_str_int(self, r: BinaryReader) -> Dict[str, int]:
        """Read Map<String, Int32>: S32 count + [{String, Int32}...]"""
        count = r.read_i32()
        if count < 0 or count > 100000:
            raise ValueError(f"Implausible map count: {count}")
        result = {}
        for _ in range(count):
            k = r.read_string()
            v = r.read_i32()
            result[k] = v
        return result

    def _read_hash128_array(self, r: BinaryReader) -> int:
        """Read Array<Hash128>: S32 count + [16B...], return count."""
        count = r.read_i32()
        if count < 0 or count > 100000:
            raise ValueError(f"Implausible hash128 array count: {count}")
        r.read(count * 16)
        return count

    # ── SerializedShaderFloatValue = float(4B) + FastPropertyName(String) ──
    # NOTE: FastPropertyName is serialized as a string in binary format, NOT as int!

    def _skip_float_value(self, r: BinaryReader):
        r.read_f32()       # val
        r.read_string()    # name (FastPropertyName → serialized as string)

    # ── SerializedShaderVectorValue = 4 × FloatValue + FastPropertyName(String) ──

    def _skip_vector_value(self, r: BinaryReader):
        for _ in range(4):
            self._skip_float_value(r)
        r.read_string()    # name (FastPropertyName → serialized as string)

    # ── SerializedStencilOp = 4 × FloatValue ──

    def _skip_stencil_op(self, r: BinaryReader):
        for _ in range(4):
            self._skip_float_value(r)

    # ── SerializedShaderRTBlendState = 7 × FloatValue ──

    def _skip_rt_blend_state(self, r: BinaryReader):
        for _ in range(7):
            self._skip_float_value(r)

    # ── SerializedTextureProperty ──

    def _read_texture_property(self, r: BinaryReader):
        r.read_string()   # m_DefaultName
        r.read_i32()       # m_TexDim (enum)

    # ── SerializedProperty ──

    def _read_property(self, r: BinaryReader):
        r.read_string()   # m_Name
        r.read_string()   # m_Description
        # m_Attributes[]: Array<String>
        self._read_string_array(r)
        r.read_i32()       # m_Type (enum)
        r.read_u32()       # m_Flags
        # m_DefValue[4]: 4 × float
        r.pos += 16
        # m_DefTexture
        self._read_texture_property(r)

    # ── SerializedProperties (m_PropInfo) ──

    def _read_prop_info(self, r: BinaryReader):
        """Read SerializedProperties: just m_Props array."""
        count = r.read_i32()
        if count < 0 or count > 10000:
            raise ValueError(f"Implausible property count: {count}")
        for _ in range(count):
            self._read_property(r)

    # ── SerializedProgramParameters ──

    def _read_vector_param(self, r: BinaryReader):
        """VectorParameter: nameIdx(i32) + idx(i32) + arrSz(i32) + type(enumByte) + dim(i8) + Align"""
        r.pos += 12  # nameIdx + idx + arrSz
        r.pos += 2   # type(1B) + dim(1B)
        r.align(4)

    def _read_matrix_param(self, r: BinaryReader):
        """MatrixParameter: nameIdx(i32) + idx(i32) + arrSz(i32) + type(enumByte) + rowCount(i8) + Align"""
        r.pos += 12
        r.pos += 2
        r.align(4)

    def _read_texture_param(self, r: BinaryReader):
        """TextureParameter: nameIdx(i32) + idx(i32) + samplerIdx(i32) + multiSampled(bool) + dim(enumByte) + Align"""
        r.pos += 12  # nameIdx + idx + samplerIdx
        r.pos += 2   # multiSampled(1B) + dim(1B)
        r.align(4)

    def _read_buffer_binding(self, r: BinaryReader):
        """BufferBinding: nameIdx(i32) + idx(i32) + arrSz(i32)"""
        r.pos += 12

    def _read_uav_param(self, r: BinaryReader):
        """UAVParameter: nameIdx(i32) + idx(i32) + origIdx(i32)"""
        r.pos += 12

    def _read_sampler_param(self, r: BinaryReader):
        """SamplerParameter: bits(u32) + bindPoint(i32)"""
        r.pos += 8

    def _read_struct_param(self, r: BinaryReader):
        """StructParameter: nameIdx + idx + arrSz + structSz + vectorMembers[] + matrixMembers[]"""
        r.pos += 16  # nameIdx(4) + idx(4) + arrSz(4) + structSz(4)
        # m_VectorMembers[]
        vc = r.read_i32()
        for _ in range(vc):
            self._read_vector_param(r)
        # m_MatrixMembers[]
        mc = r.read_i32()
        for _ in range(mc):
            self._read_matrix_param(r)

    def _read_constant_buffer(self, r: BinaryReader):
        """ConstantBuffer (Version 2): nameIdx + matrixParams[] + vectorParams[] + structParams[] + size + isPartialCB + Align"""
        r.read_i32()  # m_NameIndex
        # m_MatrixParams[]
        mc = r.read_i32()
        for _ in range(mc):
            self._read_matrix_param(r)
        # m_VectorParams[]
        vc = r.read_i32()
        for _ in range(vc):
            self._read_vector_param(r)
        # m_StructParams[]
        sc = r.read_i32()
        for _ in range(sc):
            self._read_struct_param(r)
        r.read_i32()  # m_Size
        r.read_u8()   # m_IsPartialCB (bool)
        r.align(4)

    def _read_program_parameters(self, r: BinaryReader):
        """Read SerializedProgramParameters."""
        # m_VectorParams[]
        vc = r.read_i32()
        for _ in range(vc):
            self._read_vector_param(r)
        # m_MatrixParams[]
        mc = r.read_i32()
        for _ in range(mc):
            self._read_matrix_param(r)
        # m_TextureParams[]
        tc = r.read_i32()
        for _ in range(tc):
            self._read_texture_param(r)
        # m_BufferParams[]
        bc = r.read_i32()
        for _ in range(bc):
            self._read_buffer_binding(r)
        # m_ConstantBuffers[]
        cbc = r.read_i32()
        for _ in range(cbc):
            self._read_constant_buffer(r)
        # m_ConstantBufferBindings[]
        cbbc = r.read_i32()
        for _ in range(cbbc):
            self._read_buffer_binding(r)
        # m_UAVParams[]
        uc = r.read_i32()
        for _ in range(uc):
            self._read_uav_param(r)
        # m_Samplers[]
        sc = r.read_i32()
        for _ in range(sc):
            self._read_sampler_param(r)

    # ── SerializedBindChannels ──

    def _read_bind_channels(self, r: BinaryReader):
        """SerializedBindChannels: m_Channels[] (each = 2B) + Align + m_FullChannelMask(i32)"""
        count = r.read_i32()
        if count < 0 or count > 100000:
            raise ValueError(f"Implausible bind channel count: {count}")
        r.pos += count * 2  # each SerializedBindChannel = source(1B) + target(1B)
        r.align(4)
        r.read_i32()  # m_FullChannelMask (enum)

    # ── SerializedSubProgram (旧格式, Version 5) ──

    def _read_sub_program(self, r: BinaryReader) -> SubProgramInfo:
        """Read a SerializedSubProgram (old format)."""
        blob_index = r.read_u32()  # m_BlobIndex
        r.read_u32()  # m_BlobDecompressSize (v5)
        self._read_bind_channels(r)  # m_Channels
        # m_KeywordIndices[]: Array<UInt16>
        keyword_indices = self._read_u16_array(r)
        graphics_tier = r.read_u8()  # m_GraphicsTier (enumByte)
        gpu_program_type = r.read_u8()  # m_GpuProgramType (enumByte)
        r.align(4)
        # m_Parameters (v4+, we assume v5)
        self._read_program_parameters(r)
        # m_ShaderRequirements: UInt64
        r.read_u64()

        return SubProgramInfo(
            gpu_program_type=gpu_program_type,
            keyword_indices=keyword_indices,
            blob_index=blob_index,
            source_kind="old_sub_program",
            graphics_tier=graphics_tier,
        )

    # ── SerializedPlayerSubProgram — 每个就是一个变体！ ──

    def _read_player_sub_program(self, r: BinaryReader, player_tier: int) -> SubProgramInfo:
        """Read a SerializedPlayerSubProgram = one compiled variant."""
        blob_index = r.read_u32()  # m_BlobIndex
        # m_KeywordIndices[]: Array<UInt16>
        keyword_indices = self._read_u16_array(r)
        # m_ShaderRequirements: UInt64
        r.read_u64()
        # m_GpuProgramType: enumByte
        gpu_program_type = r.read_u8()
        r.align(4)

        return SubProgramInfo(
            gpu_program_type=gpu_program_type,
            keyword_indices=keyword_indices,
            blob_index=blob_index,
            source_kind="player_sub_program",
            player_tier=player_tier,
        )

    # ── SerializedProgram ──

    def _read_program(self, r: BinaryReader) -> List[SubProgramInfo]:
        """Read a SerializedProgram, return list of all sub-programs (variants).

        Returns player sub-programs if present, otherwise old sub-programs.
        """
        # m_SubPrograms[]: Array<SerializedSubProgram> (old format)
        old_count = r.read_i32()
        if old_count < 0 or old_count > 500000:
            raise ValueError(f"Implausible sub program count: {old_count}")
        old_subs = []
        for _ in range(old_count):
            old_subs.append(self._read_sub_program(r))

        # m_PlayerSubPrograms[][]: Array<Array<SerializedPlayerSubProgram>> (3 tiers)
        tier_count = r.read_i32()
        if tier_count < 0 or tier_count > 100:
            raise ValueError(f"Implausible tier count: {tier_count}")
        player_subs = []
        for tier_index in range(tier_count):
            inner_count = r.read_i32()
            if inner_count < 0 or inner_count > 500000:
                raise ValueError(f"Implausible player sub program count: {inner_count}")
            for _ in range(inner_count):
                player_subs.append(self._read_player_sub_program(r, tier_index))

        # m_ParameterBlobIndices[][]: Array<Array<UInt32>>
        pbi_outer = r.read_i32()
        if pbi_outer < 0 or pbi_outer > 100:
            raise ValueError(f"Implausible parameter blob indices outer count: {pbi_outer}")
        for _ in range(pbi_outer):
            self._read_u32_array(r)

        # m_CommonParameters: SerializedProgramParameters
        self._read_program_parameters(r)

        # m_SerializedKeywordStateMask[]: Array<UInt16>
        self._read_u16_array(r)

        # m_UserGlobal[]: Array<String>
        self._read_string_array(r)
        # m_UserLocal[]: Array<String>
        self._read_string_array(r)
        # m_Builtin[]: Array<String>
        self._read_string_array(r)

        # Prefer player sub-programs (new format) over old
        return player_subs if player_subs else old_subs

    # ── SerializedShaderState (Version 2) ──

    def _read_shader_state(self, r: BinaryReader) -> dict:
        """Read SerializedShaderState — large fixed-ish structure.
        Returns dict with 'name' and 'tags'."""
        state_name = r.read_string()  # m_Name

        # rtBlend[8]: 8 × SerializedShaderRTBlendState (56B each)
        for _ in range(8):
            self._skip_rt_blend_state(r)

        # rtSeparateBlend: bool + Align
        r.read_u8()
        r.align(4)

        # zClip, zTest, zWrite, culling, conservative, offsetFactor, offsetUnits, alphaToMask
        # = 8 × SerializedShaderFloatValue (8B each)
        for _ in range(8):
            self._skip_float_value(r)

        # stencilOp, stencilOpFront, stencilOpBack = 3 × SerializedStencilOp (32B each)
        for _ in range(3):
            self._skip_stencil_op(r)

        # stencilReadMask, stencilWriteMask, stencilRef = 3 × FloatValue (8B each)
        for _ in range(3):
            self._skip_float_value(r)

        # fogStart, fogEnd, fogDensity = 3 × FloatValue (8B each)
        for _ in range(3):
            self._skip_float_value(r)

        # fogColor: SerializedShaderVectorValue (36B)
        self._skip_vector_value(r)

        # fogMode: Int32 (enum)
        r.read_i32()

        # gpuProgramID: Int32
        r.read_i32()

        # m_Tags: Map<String, String>
        state_tags = self._read_map_str_str(r)

        # m_LOD: Int32
        r.read_i32()

        # lighting: bool + Align
        r.read_u8()
        r.align(4)

        return {'name': state_name, 'tags': state_tags}

    # ── SerializedPass (Version 2) ──

    def _read_pass(self, r: BinaryReader) -> dict:
        """Read a SerializedPass, return dict with programs and metadata."""
        # m_EditorDataHash[]: Array<Hash128> (game release = empty)
        self._read_hash128_array(r)

        # m_Platforms[]: Array<UInt8>
        self._read_u8_array(r)

        # m_NameIndices: Map<String, Int32>
        name_indices = self._read_map_str_int(r)

        # m_Type: Int32 (enum)
        pass_type = r.read_i32()

        # m_State: SerializedShaderState
        state = self._read_shader_state(r)

        # m_ProgramMask: UInt32
        r.read_u32()

        # 6 programs: Vertex, Fragment, Geometry, Hull, Domain, RayTracing
        programs = {}
        for stage_name in SHADER_STAGE_NAMES:
            subs = self._read_program(r)
            if subs:
                programs[stage_name] = subs

        # m_HasInstancingVariant: bool (1B)
        r.read_u8()
        # m_HasProceduralInstancingVariant: bool (1B)
        r.read_u8()
        r.align(4)

        # m_UseName: String
        r.read_string()
        # m_Name: String
        pass_name = r.read_string()
        # m_TextureName: String
        r.read_string()
        # m_Tags: Map<String, String>
        pass_tags = self._read_map_str_str(r)
        # No m_PackageRequirements in game release

        # Determine best display name: LightMode tag > state tags > state name > pass name
        all_tags = {}
        all_tags.update(state.get('tags', {}))
        all_tags.update(pass_tags)
        light_mode = all_tags.get('LightMode', all_tags.get('LIGHTMODE', ''))

        display_name = light_mode or state.get('name', '') or pass_name
        return {'name': display_name, 'programs': programs, 'tags': all_tags, 'type': pass_type}

    # ── SerializedSubShader ──

    def _read_sub_shader(self, r: BinaryReader) -> dict:
        """Read a SerializedSubShader."""
        # m_Passes[]: Array<SerializedPass>
        pass_count = r.read_i32()
        if pass_count < 0 or pass_count > 1000:
            raise ValueError(f"Implausible pass count: {pass_count}")
        passes = []
        for _ in range(pass_count):
            passes.append(self._read_pass(r))

        # m_Tags: Map<String, String>
        self._read_map_str_str(r)

        # m_LOD: Int32
        lod = r.read_i32()

        # No m_PackageRequirements in game release

        return {'passes': passes, 'lod': lod}

    def _read_sub_shaders(self, r: BinaryReader) -> list:
        """Read Array<SerializedSubShader>."""
        count = r.read_i32()
        if count < 0 or count > 1000:
            raise ValueError(f"Implausible sub shader count: {count}")
        result = []
        for _ in range(count):
            result.append(self._read_sub_shader(r))
        return result

    # ── m_KeywordFlags detection ──

    def _try_read_keyword_flags(self, r: BinaryReader, keyword_count: int) -> List[int]:
        """Try to read m_KeywordFlags array (may not exist).

        m_KeywordFlags is Array<UInt8> that may not be present in older data.
        We detect this by checking if the next bytes look like a plausible
        UInt8 array whose count matches keyword_count.
        """
        if keyword_count == 0:
            return []

        saved_pos = r.pos
        try:
            count = r.read_i32()
            if count == keyword_count and count > 0:
                flags = []
                for _ in range(count):
                    flags.append(r.read_u8())
                r.align(4)
                return flags
            else:
                # Not keyword flags — rewind
                r.pos = saved_pos
                return []
        except (EOFError, ValueError):
            r.pos = saved_pos
            return []


@dataclass
class TypeTreeNode:
    version: int
    level: int
    type_flags: int
    type_str_offset: int
    name_str_offset: int
    byte_size: int
    index: int
    meta_flag: int
    ref_type_hash: int
    type_name: str = ""
    field_name: str = ""
    children: list = field(default_factory=list)


class SerializedFileParser:
    """Parse a Unity SerializedFile and extract Shader objects."""

    def __init__(self):
        self._structured_reader = StructuredShaderReader()

    def parse(self, data: bytes) -> List[ShaderInfo]:
        """Parse a SerializedFile and extract Shader objects."""
        if len(data) < 20:
            return []

        try:
            return self._parse_internal(data)
        except Exception:
            return []

    def _parse_internal(self, data: bytes) -> List[ShaderInfo]:
        reader = BinaryReader(data, big_endian=True)

        # SerializedFile header (always big-endian initially)
        metadata_size = reader.read_u32()
        file_size_32 = reader.read_u32()
        version = reader.read_u32()
        data_offset_32 = reader.read_u32()

        # Sanity check version
        if version < 5 or version > 50:
            return []

        if version >= 9:
            endianness = reader.read_u8()
            reader.read(3)  # reserved
        else:
            endianness = 0

        file_size = file_size_32
        data_offset = data_offset_32

        if version >= 22:
            # Extended header is ALWAYS big-endian (header endianness, not file endianness)
            metadata_size = reader.read_u32()  # still BE here
            file_size = reader.read_i64()
            data_offset = reader.read_i64()
            reader.read_i64()  # unknown

        # Sanity check
        if data_offset < 0 or data_offset > len(data):
            return []

        # NOW switch to file endianness for metadata
        is_big_endian = endianness != 0
        reader.big_endian = is_big_endian
        reader._endian = '>' if is_big_endian else '<'

        # Unity version string
        if version >= 7:
            unity_version = reader.read_cstring()

        if version >= 8:
            target_platform = reader.read_u32()

        # Type tree
        if version >= 13:
            enable_type_tree = reader.read_u8() != 0
        else:
            enable_type_tree = True

        type_count = reader.read_u32()
        if type_count > 10000:
            return []

        types = []
        for ti in range(type_count):
            try:
                type_info = self._read_serialized_type(reader, version, enable_type_tree)
                types.append(type_info)
            except Exception:
                # Can't parse types, bail
                return []

        # Object info table
        if version >= 7 and version < 14:
            big_id_enabled = reader.read_u32()
        else:
            big_id_enabled = 0

        object_count = reader.read_u32()
        if object_count > 100000:
            return []

        objects = []
        for _ in range(object_count):
            try:
                if version >= 14:
                    reader.align(4)
                    path_id = reader.read_i64()
                elif big_id_enabled:
                    path_id = reader.read_i64()
                else:
                    path_id = reader.read_i32()

                if version >= 22:
                    byte_start = reader.read_i64()
                else:
                    byte_start = reader.read_u32()
                byte_size = reader.read_u32()
                type_id = reader.read_u32()

                if version < 16:
                    class_id = reader.read_u16()
                else:
                    class_id = types[type_id]['class_id'] if type_id < len(types) else -1

                if version < 16:
                    reader.read_u16()  # is_destroyed

                if version >= 15 and version < 17:
                    reader.read_u8()  # stripped

                objects.append({
                    'path_id': path_id,
                    'byte_start': byte_start,
                    'byte_size': byte_size,
                    'type_id': type_id,
                    'class_id': class_id,
                })
            except Exception:
                break

        # Extract shader objects
        shaders = []
        for obj in objects:
            if obj['class_id'] != UNITY_SHADER_CLASS_ID:
                continue

            abs_offset = data_offset + obj['byte_start']
            end_offset = abs_offset + obj['byte_size']
            if abs_offset < 0 or end_offset > len(data):
                continue

            obj_data = data[abs_offset:end_offset]
            if len(obj_data) < 8:
                continue

            try:
                shader = self._structured_reader.read_shader(obj_data, is_big_endian)
                if shader and shader.name:
                    shaders.append(shader)
            except Exception:
                # Fallback to heuristic parsing
                try:
                    shader = self._extract_shader_heuristic(obj_data, is_big_endian)
                    if shader and shader.name:
                        shader.name = shader.name + " [heuristic]"
                        shaders.append(shader)
                except Exception:
                    pass

        return shaders

    def _read_serialized_type(self, reader: BinaryReader, version: int, has_type_tree: bool) -> dict:
        """Read a SerializedType entry."""
        result = {}

        if version >= 17:
            class_id = reader.read_u32()
        else:
            class_id = reader.read_i16()

        result['class_id'] = class_id

        if version >= 16:
            reader.read_u8()  # is_stripped

        if version >= 17:
            reader.read_i16()  # script_type_index

        if version >= 13:
            if (version < 16 and class_id < 0) or (version >= 16 and class_id == 114):
                reader.read(16)  # script Hash128
            reader.read(16)  # old type Hash128

        if has_type_tree:
            self._skip_type_tree(reader, version)

            # Type dependencies only present WITH type tree
            if version >= 21:
                if class_id == UNITY_SHADER_CLASS_ID:
                    dep_count = reader.read_u32()
                    reader.read(dep_count * 4)

        return result

    def _skip_type_tree(self, reader: BinaryReader, version: int):
        """Skip over a type tree blob efficiently."""
        if version >= 12:
            node_count = reader.read_u32()
            string_buffer_size = reader.read_u32()
            if node_count > 100000 or string_buffer_size > 10000000:
                raise ValueError(f"Implausible type tree: {node_count} nodes, {string_buffer_size} string bytes")
            # Each node: version(2) + level(1) + type_flags(1) + type_offset(4) + name_offset(4) +
            #            byte_size(4) + index(4) + meta_flag(4) = 24 bytes
            #            + ref_type_hash(8) if version >= 19 = 32 bytes
            node_size = 32 if version >= 19 else 24
            reader.read(node_count * node_size)
            reader.read(string_buffer_size)

    @staticmethod
    def _is_valid_shader_name(name: str) -> bool:
        """Check if a string looks like a valid shader name (not property info)."""
        if not name or len(name) > 200:
            return False
        # Must have at least one letter
        if not any(c.isalpha() for c in name):
            return False
        # Shader names should not contain control characters (tabs, newlines)
        if any(c in name for c in '\t\n\r'):
            return False
        # Shader names don't start/end with whitespace
        if name != name.strip():
            return False
        # Property description patterns — NOT shader names
        if any(p in name for p in ['OnValueChanged', 'BeginFoldoutHeader', 'ProEnum(', 'Enum(Unity',
                                     'Surface Type', 'Blend Mode', 'Face Texture', 'Face Color',
                                     'Cull Mode', 'Tint Color', 'Base Color']):
            return False
        # Long runs of spaces indicate property descriptions
        if '   ' in name:
            return False
        # Shader names should contain '/' (e.g., Hidden/Internal-StencilWrite, JNShader/T3/Foo)
        # Single-word names starting with '_' are property names (e.g., _MainTex, _Color)
        if '/' not in name:
            if name.startswith('_'):
                return False
            # Allow non-'/' names only if they look like identifiers (e.g., "Legacy Shaders")
            # But short names without '/' are suspicious
            if len(name) < 5:
                return False
        # Names containing '(' or '=' are property descriptions
        if '(' in name or '=' in name:
            return False
        return True

    def _extract_shader_heuristic(self, obj_data: bytes, is_big_endian: bool) -> Optional[ShaderInfo]:
        """Extract shader name and keywords from object data using heuristic parsing.

        Player-build Shader object layout (m_ObjectHideFlags stripped):
          m_Name (NamedObject, String) — usually empty in player builds
          m_ParsedForm (SerializedShader):
            m_PropInfo.m_Props[] — properties (variable size, need to skip)
            m_SubShaders[] — subshaders (variable size, complex nested, need to skip)
            m_KeywordNames[] — array of strings (what we want!)
            m_KeywordFlags[] — array of UInt8
            m_Name — THE ACTUAL SHADER NAME (string, comes AFTER keywords!)
            m_CustomEditorName — string
            m_FallbackName — string
            ...
        """
        endian = '>' if is_big_endian else '<'

        if len(obj_data) < 12:
            return None

        # Step 1: Find keywords (they come before the shader name in m_ParsedForm)
        # Try offset 0 (no hide flags, name_len=0 for player builds) and offset 4 (with hide flags)
        search_start = 0

        # Check if outer m_Name exists at offset 0
        outer_name_len = struct.unpack_from(f'{endian}I', obj_data, 0)[0]
        outer_name = ""
        if outer_name_len == 0:
            search_start = 4  # skip empty name, start of m_ParsedForm
        elif 0 < outer_name_len < 500:
            # Could be a real name or the start of m_PropInfo
            try:
                candidate = obj_data[4:4 + outer_name_len].decode('utf-8', errors='strict')
                if self._is_valid_shader_name(candidate):
                    outer_name = candidate
                    search_start = 4 + outer_name_len
                    search_start += (4 - search_start % 4) % 4
                else:
                    search_start = 4  # Treat as empty name (len=0 was at an earlier offset)
            except (UnicodeDecodeError, ValueError):
                search_start = 4

        # Step 2: Find keyword array in the data
        keywords = self._find_keyword_array(obj_data, search_start, endian)

        # Step 3: Find the shader name AFTER the keyword array
        # m_Name comes after m_KeywordNames and m_KeywordFlags in SerializedShader
        shader_name = outer_name
        if not shader_name and keywords:
            # Scan for the shader name after where keywords were found
            # m_KeywordNames → m_KeywordFlags(byte array) → m_Name(string)
            shader_name = self._find_shader_name_after_keywords(obj_data, search_start, endian)

        if not shader_name:
            # Fallback: scan entire object for any valid shader name
            shader_name = self._find_shader_name(obj_data, 0, endian)

        if not shader_name and not keywords:
            return None

        if not shader_name:
            shader_name = "<unknown_shader>"

        return ShaderInfo(name=shader_name, keywords=keywords)

    def _find_shader_name_after_keywords(self, data: bytes, start: int, endian: str) -> Optional[str]:
        """Find m_Name by scanning for strings with '/' after the keyword array region.

        In SerializedShader, field order is:
          m_PropInfo → m_SubShaders → m_KeywordNames → m_KeywordFlags → m_Name
        So m_Name should be found after the keyword array.
        """
        # Scan the region after 60% of the data (keywords are typically in the middle)
        # to increase chances of finding the actual m_Name and not a property name
        scan_start = start + len(data) // 4  # Start at 25% to skip properties
        return self._find_shader_name(data, scan_start, endian)

    def _find_shader_name(self, data: bytes, start: int, endian: str) -> Optional[str]:
        """Scan for the shader name (m_ParsedForm.m_Name) in the object data.

        Shader names always contain '/' (e.g. Hidden/XXX, JNShader/T3/Foo).
        They must be printable ASCII, no tabs/newlines, no property-description patterns.
        """
        end = min(len(data), start + 50000)
        pos = start
        while pos < end - 8:
            slen = struct.unpack_from(f'{endian}I', data, pos)[0]
            if 5 < slen < 200 and pos + 4 + slen <= len(data):
                try:
                    s = data[pos + 4:pos + 4 + slen].decode('utf-8', errors='strict')
                    # Strict shader name: must contain '/', printable ASCII, no control chars
                    if ('/' in s
                            and s == s.strip()
                            and all(0x20 <= ord(c) < 0x7F for c in s)
                            and '(' not in s and '=' not in s
                            and '   ' not in s
                            and '\t' not in s
                            and not s.startswith('$')
                            and not s.startswith('_')):
                        return s
                except (UnicodeDecodeError, ValueError):
                    pass
            pos += 4
        return None

    def _find_keyword_array(self, data: bytes, start: int, endian: str) -> List[str]:
        """Scan for a keyword string array pattern in shader object data."""
        end = min(len(data), start + 10000)  # Keywords are near the start
        pos = start

        import re
        # Shader keywords are UPPER_CASE_WITH_UNDERSCORES, optionally prefixed with _
        # e.g. "STEREO_INSTANCING_ON", "_MAIN_LIGHT_SHADOWS", "DISABLE_REVERSED_Z"
        # Property names are mixed-case: "_SpecColor", "_MainTex", "_Color"
        _KW_RE = re.compile(r'^_?[A-Z][A-Z0-9_]*$')

        best_keywords = []
        best_pos = -1

        while pos < end - 8:
            count = struct.unpack_from(f'{endian}I', data, pos)[0]

            # Plausible keyword count: 1-5000
            if count < 1 or count > 5000 or pos + 4 + count * 2 > len(data):
                pos += 4
                continue

            # Quick validation: check first string length looks reasonable
            first_slen = struct.unpack_from(f'{endian}I', data, pos + 4)[0]
            if first_slen < 2 or first_slen > 200:
                pos += 4
                continue

            try:
                sample_count = min(count, 8)
                keywords, valid = self._try_read_string_array(data, pos + 4, sample_count, endian)
                if valid and len(keywords) >= min(3, count):
                    # Strict check: ALL sampled strings must be UPPER_CASE keyword pattern
                    keyword_like = sum(1 for kw in keywords if kw and _KW_RE.match(kw))
                    if keyword_like == len(keywords):
                        # Read all keywords
                        all_kw, _ = self._try_read_string_array(data, pos + 4, count, endian)
                        # Keep the largest matching array (m_KeywordNames is the big one)
                        if len(all_kw) > len(best_keywords):
                            best_keywords = all_kw
                            best_pos = pos
            except Exception:
                pass

            pos += 4

        return [kw for kw in best_keywords if kw]  # Filter empty strings

        return []

    def _try_read_string_array(self, data: bytes, pos: int, count: int, endian: str) -> Tuple[List[str], bool]:
        """Try to read an array of length-prefixed strings."""
        strings = []
        for _ in range(count):
            if pos + 4 > len(data):
                return strings, False
            slen = struct.unpack_from(f'{endian}I', data, pos)[0]
            pos += 4
            if slen > 500 or pos + slen > len(data):
                return strings, False
            try:
                s = data[pos:pos + slen].decode('utf-8')
                # Basic validity: should be printable ASCII
                if not all(0x20 <= ord(c) < 0x7F or c in '\t\n' for c in s):
                    return strings, False
            except UnicodeDecodeError:
                return strings, False
            strings.append(s)
            pos += slen
            pos += (4 - pos % 4) % 4  # align to 4
        return strings, True


# ════════════════════════════════════════════════════════════════════
# Material Keyword Scanner
# ════════════════════════════════════════════════════════════════════

import re as _re

@dataclass
class MaterialScanResult:
    """Result of scanning .mat files for keyword usage."""
    # {keyword: count_of_materials_using_it}
    keyword_material_count: Dict[str, int] = field(default_factory=lambda: defaultdict(int))
    # {keyword: [list_of_mat_file_paths]}
    keyword_materials: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))
    # All keywords found across all materials
    all_keywords: set = field(default_factory=set)
    # Total materials scanned / parsed
    total_scanned: int = 0
    total_parsed: int = 0


class MaterialKeywordScanner:
    """Scan Unity .mat files to extract enabled shader keywords.

    Supports two YAML formats:
    - New: m_ValidKeywords array (- KEYWORD lines)
    - Old: m_ShaderKeywords: "KEYWORD1 KEYWORD2 ..."
    """

    # Regex for m_ValidKeywords array items
    _RE_VALID_KW_ITEM = _re.compile(r'^\s+-\s+(\S+)')
    # Regex for m_ShaderKeywords string value
    _RE_SHADER_KW = _re.compile(r'm_ShaderKeywords:\s*(.*)')
    # Regex for m_ValidKeywords: section start
    _RE_VALID_KW_START = _re.compile(r'm_ValidKeywords:')
    # Keyword pattern: UPPER_CASE identifiers (shader keywords)
    _RE_KW_PATTERN = _re.compile(r'^_?[A-Z][A-Z0-9_]*$')

    def scan_materials(self, project_dir: str) -> MaterialScanResult:
        """Scan Assets/ under project_dir for .mat files, extract keywords."""
        result = MaterialScanResult()
        assets_dir = os.path.join(project_dir, "Assets")
        if not os.path.isdir(assets_dir):
            # Try project_dir itself
            assets_dir = project_dir

        mat_files = []
        for root, dirs, files in os.walk(assets_dir):
            for f in files:
                if f.endswith('.mat'):
                    mat_files.append(os.path.join(root, f))

        result.total_scanned = len(mat_files)
        print(f"  扫描到 {len(mat_files)} 个 .mat 文件")

        for mat_path in mat_files:
            try:
                keywords = self._parse_mat_file(mat_path)
                if keywords is not None:
                    result.total_parsed += 1
                    rel_path = os.path.relpath(mat_path, project_dir)
                    for kw in keywords:
                        if kw and self._RE_KW_PATTERN.match(kw):
                            result.keyword_material_count[kw] += 1
                            result.keyword_materials[kw].append(rel_path)
                            result.all_keywords.add(kw)
            except Exception:
                pass

        print(f"  成功解析 {result.total_parsed} 个材质文件")
        print(f"  发现 {len(result.all_keywords)} 个不同的 keyword")
        return result

    def _parse_mat_file(self, mat_path: str) -> Optional[set]:
        """Parse a single .mat YAML file, return set of enabled keywords."""
        try:
            with open(mat_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read(64 * 1024)  # Cap at 64KB per file
        except (OSError, IOError):
            return None

        keywords = set()

        # Strategy 1: m_ValidKeywords array
        in_valid_kw = False
        for line in content.split('\n'):
            if self._RE_VALID_KW_START.search(line):
                in_valid_kw = True
                continue
            if in_valid_kw:
                m = self._RE_VALID_KW_ITEM.match(line)
                if m:
                    keywords.add(m.group(1))
                else:
                    # End of array (line doesn't match "- ITEM" pattern)
                    in_valid_kw = False

            # Strategy 2: m_ShaderKeywords string
            m = self._RE_SHADER_KW.match(line.strip())
            if m:
                kw_str = m.group(1).strip()
                if kw_str:
                    for kw in kw_str.split():
                        keywords.add(kw)

        return keywords if keywords else set()


# ════════════════════════════════════════════════════════════════════
# Report Generator
# ════════════════════════════════════════════════════════════════════

def generate_report(shaders: List[ShaderInfo], output_dir: str, source_label: str = "",
                    material_scan: Optional[MaterialScanResult] = None, export_xlsx: bool = True):
    """Generate Markdown analysis reports and Excel unless disabled."""
    os.makedirs(output_dir, exist_ok=True)

    # Deduplicate shaders by name — keep the one with more variants (or more keywords)
    dedup = {}
    for s in shaders:
        existing = dedup.get(s.name)
        if existing is None or s.total_variants > existing.total_variants or (
                s.total_variants == existing.total_variants and len(s.keywords) > len(existing.keywords)):
            dedup[s.name] = s
    shaders = list(dedup.values())

    total_variants = sum(s.total_variants for s in shaders)
    total_unique_variants = sum(s.total_unique_variants for s in shaders)
    total_keywords = sum(len(s.keywords) for s in shaders)
    structured_count = sum(1 for s in shaders if not s.name.endswith("[heuristic]"))
    heuristic_count = sum(1 for s in shaders if s.name.endswith("[heuristic]"))

    # Sort by variant count
    sorted_shaders = sorted(shaders, key=lambda s: s.total_variants, reverse=True)

    # ── summary.md ──
    sb = []
    sb.append("# Shader Variant Analysis — Package Parser Report")
    sb.append("")
    if source_label:
        sb.append(f"**数据来源**: {source_label}")
        sb.append("")
    sb.append(f"**Shader 总数**: {len(shaders)}")
    sb.append(f"**总原始变体数**: {total_variants:,}")
    sb.append(f"**总唯一变体数**: {total_unique_variants:,}")
    sb.append(f"**总 Keyword 数**: {total_keywords:,}")
    sb.append(f"**结构化解析成功**: {structured_count}  |  **启发式 fallback**: {heuristic_count}")
    sb.append("")

    sb.append("## Shader 列表（按变体数排序）")
    sb.append("")
    sb.append("| # | Shader | Keywords | 原始变体数 | 唯一变体数 | SubShaders | Passes |")
    sb.append("|---|--------|----------|------------|------------|------------|--------|")
    for i, s in enumerate(sorted_shaders, 1):
        total_passes = sum(len(ss.passes) for ss in s.sub_shaders)
        sb.append(f"| {i} | `{s.name}` | {len(s.keywords)} | {s.total_variants:,} | {s.total_unique_variants:,} | {len(s.sub_shaders)} | {total_passes} |")
    sb.append("")

    # Top 20 variant-heavy shaders
    top20 = [s for s in sorted_shaders if s.total_variants > 0][:20]
    if top20:
        sb.append("## Top 20 变体最多的 Shader")
        sb.append("")
        sb.append("| # | Shader | 原始变体数 | 唯一变体数 | Keywords | 理论最大变体数 (2^N) |")
        sb.append("|---|--------|------------|------------|----------|---------------------|")
        for i, s in enumerate(top20, 1):
            theoretical_max = 2 ** len(s.keywords) if len(s.keywords) <= 30 else float('inf')
            ratio_str = f"{s.total_variants / theoretical_max * 100:.1f}%" if theoretical_max != float('inf') else "N/A"
            sb.append(f"| {i} | `{s.name}` | {s.total_variants:,} | {s.total_unique_variants:,} | {len(s.keywords)} | {theoretical_max:,} ({ratio_str}) |")
        sb.append("")

    with open(os.path.join(output_dir, "summary.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))

    # ── keyword_analysis.md ──
    sb = []
    sb.append("# Shader Keyword Analysis")
    sb.append("")

    # Global keyword frequency
    keyword_freq = defaultdict(int)
    keyword_shaders = defaultdict(list)
    for s in shaders:
        for kw in s.keywords:
            keyword_freq[kw] += 1
            keyword_shaders[kw].append(s.name)

    sb.append("## 全局 Keyword 频率（出现在多少个 Shader 中）")
    sb.append("")
    sb.append("| Keyword | Shader Count | Shaders |")
    sb.append("|---------|-------------|---------|")
    for kw, count in sorted(keyword_freq.items(), key=lambda x: -x[1]):
        shader_list = ", ".join(f"`{n}`" for n in keyword_shaders[kw][:5])
        if len(keyword_shaders[kw]) > 5:
            shader_list += f" ... (+{len(keyword_shaders[kw]) - 5})"
        sb.append(f"| `{kw}` | {count} | {shader_list} |")
    sb.append("")

    # Keyword ratio analysis per (shader, pass, stage)
    sb.append("## Keyword Ratio 分析（每个 Keyword 在变体中的出现率）")
    sb.append("")
    sb.append("> ratio ≈ 0.5 → keyword 没有被 strip，每个变体都有 on/off 两版（优化目标）")
    sb.append("> ratio ≈ 0 或 ≈ 1 → keyword 已被有效 strip 或总是启用")
    sb.append("")

    # Collect keyword ratios across all shaders
    ratio_entries = []  # (shader_name, pass_idx, stage, keyword, ratio, on_count, total)
    for s in shaders:
        if not s.sub_shaders or not s.keywords:
            continue
        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                for stage, subs in p.programs.items():
                    if not subs:
                        continue
                    total = len(subs)
                    # Count how many variants have each keyword ON
                    kw_on_counts = defaultdict(int)
                    for sp in subs:
                        for ki in sp.keyword_indices:
                            if ki < len(s.keywords):
                                kw_on_counts[s.keywords[ki]] += 1
                    for kw in s.keywords:
                        on_count = kw_on_counts.get(kw, 0)
                        ratio = on_count / total if total > 0 else 0.0
                        ratio_entries.append((s.name, pi, stage, kw, ratio, on_count, total))

    # Filter to ratio ≈ 0.5 (between 0.3 and 0.7) — these are the optimization targets
    strip_targets = [e for e in ratio_entries if 0.3 <= e[4] <= 0.7]
    if strip_targets:
        sb.append("### 优化目标：ratio ≈ 0.5 的 Keywords（可考虑 strip）")
        sb.append("")
        sb.append("| Shader | Pass | Stage | Keyword | Ratio | ON/Total |")
        sb.append("|--------|------|-------|---------|-------|----------|")
        # Sort by ratio closest to 0.5
        strip_targets.sort(key=lambda e: abs(e[4] - 0.5))
        for name, pi, stage, kw, ratio, on_count, total in strip_targets[:100]:
            sb.append(f"| `{name}` | {pi} | {stage} | `{kw}` | {ratio:.3f} | {on_count}/{total} |")
        if len(strip_targets) > 100:
            sb.append(f"| ... | | | | | (+{len(strip_targets) - 100} more) |")
        sb.append("")

    # Show all keyword ratios grouped by shader (for top 10 variant-heavy shaders)
    top_variant_shaders = [s for s in sorted_shaders if s.total_variants > 0][:10]
    for s in top_variant_shaders:
        if not s.sub_shaders or not s.keywords:
            continue
        sb.append(f"### `{s.name}` — Keyword Ratios")
        sb.append("")
        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                for stage, subs in p.programs.items():
                    if not subs:
                        continue
                    total = len(subs)
                    kw_on_counts = defaultdict(int)
                    for sp in subs:
                        for ki in sp.keyword_indices:
                            if ki < len(s.keywords):
                                kw_on_counts[s.keywords[ki]] += 1

                    sb.append(f"**Pass {pi} / {stage}** ({total} variants):")
                    sb.append("")
                    sb.append("| Keyword | Ratio | ON/Total | Status |")
                    sb.append("|---------|-------|----------|--------|")
                    for kw in s.keywords:
                        on_count = kw_on_counts.get(kw, 0)
                        ratio = on_count / total if total > 0 else 0.0
                        if ratio == 0.0:
                            status = "never ON"
                        elif ratio == 1.0:
                            status = "always ON"
                        elif 0.4 <= ratio <= 0.6:
                            status = "**strip target**"
                        else:
                            status = "partially stripped"
                        sb.append(f"| `{kw}` | {ratio:.3f} | {on_count}/{total} | {status} |")
                    sb.append("")

    with open(os.path.join(output_dir, "keyword_analysis.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))

    # ── shader_details.md ──
    sb = []
    sb.append("# Shader Details")
    sb.append("")
    for s in sorted_shaders:
        sb.append(f"## `{s.name}`")
        sb.append("")
        sb.append(f"**总原始变体数**: {s.total_variants:,}")
        sb.append(f"**总唯一变体数**: {s.total_unique_variants:,}")
        sb.append("")
        if s.keywords:
            sb.append(f"**Keywords ({len(s.keywords)}):**")
            sb.append("")
            for j, kw in enumerate(s.keywords):
                flag_str = ""
                if j < len(s.keyword_flags):
                    flag_str = f" (flag={s.keyword_flags[j]})"
                sb.append(f"- `{kw}`{flag_str}")
            sb.append("")

        if s.sub_shaders:
            for si, ss in enumerate(s.sub_shaders):
                sb.append(f"### SubShader {si} (LOD={ss.lod})")
                sb.append("")
                for pi, p in enumerate(ss.passes):
                    sb.append(f"#### Pass {pi}: {p.name or '(unnamed)'}")
                    sb.append("")
                    for stage, subs in p.programs.items():
                        sb.append(f"**{stage}**: {len(subs)} 原始变体 / {count_unique_subprograms(subs)} 唯一变体")
                        sb.append("")
                        # Show up to 20 variants with keyword combos
                        for vi, sp in enumerate(subs[:20]):
                            kw_names = [s.keywords[ki] if ki < len(s.keywords) else f"?{ki}" for ki in sp.keyword_indices]
                            kw_str = ", ".join(kw_names) if kw_names else "(no keywords)"
                            sb.append(
                                f"- [{vi}] GPU={sp.gpu_program_type}, Blob={sp.blob_index}, "
                                f"PlayerTier={sp.player_tier}, GraphicsTier={sp.graphics_tier}: `{kw_str}`"
                            )
                        if len(subs) > 20:
                            sb.append(f"- ... (+{len(subs) - 20} more variants)")
                        sb.append("")
        sb.append("---")
        sb.append("")

    with open(os.path.join(output_dir, "shader_details.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))

    # ── raw_data.json ──
    raw = {
        'source': source_label,
        'shader_count': len(shaders),
        'total_variants': total_variants,
        'total_unique_variants': total_unique_variants,
        'shaders': []
    }
    for s in sorted_shaders:
        shader_data = {
            'name': s.name,
            'keywords': s.keywords,
            'keyword_flags': s.keyword_flags,
            'total_variants': s.total_variants,
            'total_unique_variants': s.total_unique_variants,
            'sub_shaders': []
        }
        for ss in s.sub_shaders:
            ss_data = {'lod': ss.lod, 'passes': []}
            for p in ss.passes:
                p_data = {'name': p.name, 'programs': {}}
                for stage, subs in p.programs.items():
                    p_data['programs'][stage] = [
                        {
                            'gpu_program_type': sp.gpu_program_type,
                            'keyword_indices': sp.keyword_indices,
                            'blob_index': sp.blob_index,
                            'source_kind': sp.source_kind,
                            'player_tier': sp.player_tier,
                            'graphics_tier': sp.graphics_tier,
                        }
                        for sp in subs
                    ]
                    p_data[f'{stage}_raw_count'] = len(subs)
                    p_data[f'{stage}_unique_count'] = count_unique_subprograms(subs)
                ss_data['passes'].append(p_data)
            shader_data['sub_shaders'].append(ss_data)
        raw['shaders'].append(shader_data)

    with open(os.path.join(output_dir, "raw_data.json"), 'w', encoding='utf-8') as f:
        json.dump(raw, f, indent=2, ensure_ascii=False)

    # ── optimization_recommendations.md ──
    _generate_optimization_report(sorted_shaders, output_dir, ratio_entries, material_scan)

    # ── material_keyword_usage.md ──
    if material_scan and material_scan.all_keywords:
        _generate_material_keyword_report(sorted_shaders, output_dir, material_scan)

    # ── shader_pass_details.md ──
    _generate_shader_pass_detail_report(sorted_shaders, output_dir, material_scan)

    # ── Excel export ──
    if export_xlsx:
        _generate_xlsx_report(sorted_shaders, output_dir, ratio_entries, material_scan)

    print(f"\n报告已生成到: {output_dir}/")
    print(f"  - summary.md")
    print(f"  - keyword_analysis.md")
    print(f"  - shader_details.md")
    print(f"  - raw_data.json")
    print(f"  - optimization_recommendations.md")
    if material_scan and material_scan.all_keywords:
        print(f"  - material_keyword_usage.md")
    print(f"  - shader_pass_details.md")
    if export_xlsx:
        print(f"  - shader_variant_report.xlsx")


def _get_keyword_combo_usage(kw_names: List[str], material_scan: Optional[MaterialScanResult]) -> Tuple[bool, int]:
    """Check if a keyword combination is used by any material.

    Returns (is_used, min_material_count).
    A combo is considered 'used' if ALL keywords in it appear in at least one material.
    min_material_count is the minimum usage count among the combo's keywords.
    """
    if not material_scan or not material_scan.all_keywords:
        return (False, -1)  # -1 means no material data
    if not kw_names:
        return (True, -1)  # no-keyword variant is always "used"
    counts = []
    for kw in kw_names:
        c = material_scan.keyword_material_count.get(kw, 0)
        counts.append(c)
    min_count = min(counts) if counts else 0
    return (min_count > 0, min_count)


def _generate_shader_pass_detail_report(sorted_shaders: List[ShaderInfo], output_dir: str,
                                         material_scan: Optional[MaterialScanResult]):
    """Generate shader_pass_details.md with per-shader per-pass variant breakdown."""
    sb = []
    sb.append("# Shader-Pass 变体详情")
    sb.append("")
    if material_scan and material_scan.all_keywords:
        sb.append("> 材质使用状态基于材质 keyword 扫描结果（✓ = keyword 组合中所有 keyword 均在材质中出现）")
    else:
        sb.append("> 未提供材质扫描数据，材质使用列显示 N/A")
    sb.append("")

    for s in sorted_shaders:
        if not s.sub_shaders:
            continue

        sb.append(f"## `{s.name}` (总原始变体: {s.total_variants:,}, 总唯一变体: {s.total_unique_variants:,})")
        sb.append("")

        # Pass overview table
        sb.append("### 概览")
        sb.append("")
        stage_names_all = set()
        for ss in s.sub_shaders:
            for p in ss.passes:
                stage_names_all.update(p.programs.keys())
        stage_list = sorted(stage_names_all)

        header = "| SubShader | Pass | Pass名称 |"
        sep = "|-----------|------|----------|"
        for st in stage_list:
            header += f" {st}原始 | {st}唯一 |"
            sep += "--------|--------|"
        header += " 合计原始 | 合计唯一 |"
        sep += "----------|----------|"
        sb.append(header)
        sb.append(sep)

        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                row = f"| {si} | {pi} | {p.name or '(unnamed)'} |"
                pass_total = 0
                pass_unique_total = 0
                for st in stage_list:
                    subs = p.programs.get(st, [])
                    cnt = len(subs)
                    unique_cnt = count_unique_subprograms(subs)
                    pass_total += cnt
                    pass_unique_total += unique_cnt
                    row += f" {cnt:,} | {unique_cnt:,} |"
                row += f" {pass_total:,} | {pass_unique_total:,} |"
                sb.append(row)
        sb.append("")

        # Detailed keyword combos per pass per stage
        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                for stage, subs in p.programs.items():
                    if not subs:
                        continue

                    # Group variants by keyword combo
                    combo_counts = defaultdict(int)  # frozenset(keyword_indices) -> count
                    combo_unique_keys = defaultdict(set)
                    for sp in subs:
                        key = frozenset(sp.keyword_indices)
                        combo_counts[key] += 1
                        combo_unique_keys[key].add(subprogram_identity(sp))

                    # Sort by count descending
                    sorted_combos = sorted(combo_counts.items(), key=lambda x: -x[1])

                    sb.append(
                        f"### Pass {pi}: {p.name or '(unnamed)'} / {stage} "
                        f"({len(subs):,} 原始变体, {count_unique_subprograms(subs):,} 唯一变体, {len(sorted_combos):,} 组合)"
                    )
                    sb.append("")
                    sb.append("| # | Keyword 组合 | 原始变体数 | 唯一变体数 | 材质使用 |")
                    sb.append("|---|-------------|------------|------------|---------|")

                    for ci, (ki_set, count) in enumerate(sorted_combos, 1):
                        kw_names = sorted([s.keywords[ki] if ki < len(s.keywords) else f"?{ki}" for ki in ki_set])
                        kw_str = ", ".join(kw_names) if kw_names else "(no keywords)"
                        unique_count = len(combo_unique_keys[ki_set])

                        used, mat_count = _get_keyword_combo_usage(kw_names, material_scan)
                        if mat_count == -1:
                            usage_str = "N/A"
                        elif used:
                            usage_str = f"✓ ({mat_count}个材质)"
                        else:
                            usage_str = "✗ 无材质使用"

                        sb.append(f"| {ci} | `{kw_str}` | {count} | {unique_count} | {usage_str} |")

                    sb.append("")

        sb.append("---")
        sb.append("")

    with open(os.path.join(output_dir, "shader_pass_details.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))


def _generate_xlsx_report(sorted_shaders: List[ShaderInfo], output_dir: str,
                           ratio_entries: list, material_scan: Optional[MaterialScanResult]):
    """Generate shader_variant_report.xlsx with multiple sheets."""
    if openpyxl is None:
        print("[WARN] openpyxl 未安装，跳过 Excel 导出。安装: pip install openpyxl", file=sys.stderr)
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: 概览 ──
    ws = wb.active
    ws.title = "概览"
    headers = ["#", "Shader", "Keywords数", "变体数", "SubShaders数", "Passes数"]
    ws.append(headers)
    for i, s in enumerate(sorted_shaders, 1):
        total_passes = sum(len(ss.passes) for ss in s.sub_shaders)
        ws.append([i, s.name, len(s.keywords), s.total_variants, len(s.sub_shaders), total_passes])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── Sheet 2: Shader-Pass ──
    ws2 = wb.create_sheet("Shader-Pass")
    ws2.append(["Shader", "SubShader", "Pass", "Pass名称", "Vertex", "Fragment", "Geometry", "Hull", "Domain", "RayTracing", "合计"])
    all_stages = ["Vertex", "Fragment", "Geometry", "Hull", "Domain", "RayTracing"]
    for s in sorted_shaders:
        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                row = [s.name, si, pi, p.name or "(unnamed)"]
                pass_total = 0
                for st in all_stages:
                    cnt = len(p.programs.get(st, []))
                    row.append(cnt)
                    pass_total += cnt
                row.append(pass_total)
                ws2.append(row)
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Keyword组合 ──
    ws3 = wb.create_sheet("Keyword组合")
    ws3.append(["Shader", "SubShader", "Pass", "Pass名称", "Stage", "Keyword组合", "变体数", "材质使用", "材质数"])
    for s in sorted_shaders:
        for si, ss in enumerate(s.sub_shaders):
            for pi, p in enumerate(ss.passes):
                for stage, subs in p.programs.items():
                    if not subs:
                        continue
                    combo_counts = defaultdict(int)
                    for sp in subs:
                        key = frozenset(sp.keyword_indices)
                        combo_counts[key] += 1
                    for ki_set, count in sorted(combo_counts.items(), key=lambda x: -x[1]):
                        kw_names = sorted([s.keywords[ki] if ki < len(s.keywords) else f"?{ki}" for ki in ki_set])
                        kw_str = ", ".join(kw_names) if kw_names else "(no keywords)"
                        used, mat_count = _get_keyword_combo_usage(kw_names, material_scan)
                        if mat_count == -1:
                            usage_str = "N/A"
                            mat_count_val = ""
                        elif used:
                            usage_str = "是"
                            mat_count_val = mat_count
                        else:
                            usage_str = "否"
                            mat_count_val = 0
                        ws3.append([s.name, si, pi, p.name or "(unnamed)", stage, kw_str, count, usage_str, mat_count_val])
    ws3.auto_filter.ref = ws3.dimensions
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Keyword分析 ──
    ws4 = wb.create_sheet("Keyword分析")
    ws4.append(["Shader", "Pass", "Stage", "Keyword", "Ratio", "ON数", "Total", "Status"])
    for name, pi, stage, kw, ratio, on_count, total in ratio_entries:
        if ratio == 0.0:
            status = "never ON"
        elif ratio == 1.0:
            status = "always ON"
        elif 0.4 <= ratio <= 0.6:
            status = "strip target"
        else:
            status = "partially stripped"
        ws4.append([name, pi, stage, kw, round(ratio, 4), on_count, total, status])
    ws4.auto_filter.ref = ws4.dimensions
    ws4.freeze_panes = "A2"

    # ── Sheet 5: 材质使用 ──
    if material_scan and material_scan.all_keywords:
        ws5 = wb.create_sheet("材质使用")
        all_shader_keywords = set()
        for s in sorted_shaders:
            all_shader_keywords.update(s.keywords)
        ws5.append(["Keyword", "材质使用数", "存在于Shader"])
        for kw in sorted(material_scan.all_keywords, key=lambda k: -material_scan.keyword_material_count.get(k, 0)):
            count = material_scan.keyword_material_count.get(kw, 0)
            in_shader = "是" if kw in all_shader_keywords else "否"
            ws5.append([kw, count, in_shader])
        ws5.auto_filter.ref = ws5.dimensions
        ws5.freeze_panes = "A2"

    # ── Sheet 6: 优化建议 ──
    ws6 = wb.create_sheet("优化建议")
    ws6.append(["优先级", "Shader", "Keyword", "Avg Ratio", "影响变体数", "材质使用数", "建议"])

    # Re-compute P0/P1 candidates for xlsx
    shader_kw_ratios = defaultdict(lambda: defaultdict(dict))
    for name, pi, stage, kw, ratio, on_count, total in ratio_entries:
        shader_kw_ratios[name][kw][(pi, stage)] = (ratio, on_count, total)

    shader_kw_avg = {}
    for name, kw_dict in shader_kw_ratios.items():
        for kw, locations in kw_dict.items():
            ratios = [r for r, _, _ in locations.values()]
            totals = [t for _, _, t in locations.values()]
            avg_r = sum(ratios) / len(ratios) if ratios else 0
            total_affected = sum(totals)
            shader_kw_avg[(name, kw)] = (avg_r, total_affected)

    mat_keywords = material_scan.all_keywords if material_scan else set()

    # P0
    for (name, kw), (avg_ratio, total_affected) in sorted(shader_kw_avg.items(), key=lambda x: -x[1][1]):
        if 0.3 <= avg_ratio <= 0.7:
            mat_usage = material_scan.keyword_material_count.get(kw, 0) if material_scan else -1
            if mat_usage <= 5:
                ws6.append(["P0", name, kw, round(avg_ratio, 4), total_affected,
                            mat_usage if mat_usage >= 0 else "N/A", "改为 shader_feature"])

    # P1
    for s in sorted_shaders:
        if not s.keywords:
            continue
        for kw in s.keywords:
            locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
            if not locations:
                continue
            all_zero = all(r == 0.0 for r, _, _ in locations.values())
            if all_zero:
                total_affected = sum(t for _, _, t in locations.values())
                ws6.append(["P1", s.name, kw, 0, total_affected, "", "移除声明"])

    ws6.auto_filter.ref = ws6.dimensions
    ws6.freeze_panes = "A2"

    # Adjust column widths for readability
    for ws_item in wb.worksheets:
        for col_idx in range(1, ws_item.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws_item.iter_rows(min_col=col_idx, max_col=col_idx, max_row=min(50, ws_item.max_row)):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws_item.column_dimensions[col_letter].width = min(max_len + 2, 60)

    xlsx_path = os.path.join(output_dir, "shader_variant_report.xlsx")
    wb.save(xlsx_path)
    print(f"  - shader_variant_report.xlsx")


def _generate_optimization_report(sorted_shaders: List[ShaderInfo], output_dir: str,
                                   ratio_entries: list, material_scan: Optional[MaterialScanResult]):
    """Generate optimization_recommendations.md with P0-P5 graded suggestions."""
    mat_keywords = material_scan.all_keywords if material_scan else set()

    # Pre-compute per-shader keyword ratio summaries
    # {shader_name: {keyword: {(pass, stage): (ratio, on_count, total)}}}
    shader_kw_ratios = defaultdict(lambda: defaultdict(dict))
    for name, pi, stage, kw, ratio, on_count, total in ratio_entries:
        shader_kw_ratios[name][kw][(pi, stage)] = (ratio, on_count, total)

    # Aggregate: per (shader, keyword), compute average ratio across all pass/stages
    shader_kw_avg = {}  # {(shader, keyword): (avg_ratio, total_variants_affected)}
    for name, kw_dict in shader_kw_ratios.items():
        for kw, locations in kw_dict.items():
            ratios = [r for r, _, _ in locations.values()]
            totals = [t for _, _, t in locations.values()]
            avg_r = sum(ratios) / len(ratios) if ratios else 0
            total_affected = sum(totals)
            shader_kw_avg[(name, kw)] = (avg_r, total_affected)

    # All shader keywords set
    all_shader_keywords = set()
    for s in sorted_shaders:
        all_shader_keywords.update(s.keywords)

    sb = []
    sb.append("# Shader 变体优化建议报告")
    sb.append("")
    sb.append("> 基于包体变体数据" + ("+ 材质 keyword 扫描" if material_scan else "（无材质扫描数据）") + "的分级优化建议")
    sb.append("")

    # ── P0: multi_compile → shader_feature candidates ──
    sb.append("## P0 — `multi_compile` → `shader_feature` 候选")
    sb.append("")
    sb.append("> 条件：ratio ≈ 0.5（变体未被 strip）且 keyword 不在材质使用集合中或使用率极低")
    sb.append("> 影响：直接减半相关 keyword 的变体数量")
    sb.append("")

    p0_candidates = []
    for (name, kw), (avg_ratio, total_affected) in shader_kw_avg.items():
        if 0.3 <= avg_ratio <= 0.7:
            mat_usage = material_scan.keyword_material_count.get(kw, 0) if material_scan else -1
            # If no material scan, still flag ratio≈0.5 keywords
            if mat_usage <= 5:  # keyword unused or rarely used in materials
                p0_candidates.append((name, kw, avg_ratio, total_affected, mat_usage))

    p0_candidates.sort(key=lambda x: -x[3])  # Sort by affected variants

    if p0_candidates:
        sb.append("| Shader | Keyword | Avg Ratio | 影响变体数 | 材质使用数 | 建议 |")
        sb.append("|--------|---------|-----------|-----------|-----------|------|")
        for name, kw, avg_r, total, mat_usage in p0_candidates[:50]:
            mat_str = str(mat_usage) if mat_usage >= 0 else "N/A"
            sb.append(f"| `{name}` | `{kw}` | {avg_r:.3f} | {total:,} | {mat_str} | 改为 `shader_feature`/`shader_feature_local` |")
        if len(p0_candidates) > 50:
            sb.append(f"| ... | | | | | (+{len(p0_candidates) - 50} more) |")
    else:
        sb.append("*未发现符合条件的候选*")
    sb.append("")

    # ── P1: Never-used keywords (ratio=0 everywhere) ──
    sb.append("## P1 — 从未使用的 Keyword（所有 pass/stage ratio = 0）")
    sb.append("")
    sb.append("> 这些 keyword 被 shader 声明但编译变体中从未出现 ON 状态")
    sb.append("> 可从 shader 源码中移除声明")
    sb.append("")

    p1_candidates = []
    for s in sorted_shaders:
        if not s.keywords:
            continue
        for kw in s.keywords:
            locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
            if not locations:
                continue
            all_zero = all(r == 0.0 for r, _, _ in locations.values())
            if all_zero:
                total_affected = sum(t for _, _, t in locations.values())
                p1_candidates.append((s.name, kw, total_affected))

    p1_candidates.sort(key=lambda x: -x[2])

    if p1_candidates:
        sb.append("| Shader | Keyword | 影响变体数 | 建议 |")
        sb.append("|--------|---------|-----------|------|")
        # Highlight VFX_USE_* in non-VFX shaders
        for name, kw, total in p1_candidates[:80]:
            note = ""
            if kw.startswith("VFX_USE_") and "VFX" not in name.upper():
                note = " ⚠️ VFX keyword 在非 VFX shader 中"
            sb.append(f"| `{name}` | `{kw}` | {total:,} | 移除声明{note} |")
        if len(p1_candidates) > 80:
            sb.append(f"| ... | | | (+{len(p1_candidates) - 80} more) |")
    else:
        sb.append("*未发现从未使用的 keyword*")
    sb.append("")

    # ── P2: VFX keyword isolation ──
    sb.append("## P2 — VFX Keyword 隔离建议")
    sb.append("")
    sb.append("> `Environment/Lit` 等 shader 包含大量 VFX_USE_* keyword，但大部分 ratio = 0")
    sb.append("> 建议：VFX pass 使用 `shader_feature_local`，或拆为独立 subshader")
    sb.append("")

    p2_candidates = []
    for s in sorted_shaders:
        vfx_kws = [kw for kw in s.keywords if kw.startswith("VFX_USE_") or kw.startswith("VFX_")]
        if len(vfx_kws) < 5:
            continue
        # Count how many have non-zero ratio
        active_vfx = 0
        for kw in vfx_kws:
            locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
            if any(r > 0 for r, _, _ in locations.values()):
                active_vfx += 1
        p2_candidates.append((s.name, len(vfx_kws), active_vfx, s.total_variants))

    p2_candidates.sort(key=lambda x: -x[1])

    if p2_candidates:
        sb.append("| Shader | VFX Keywords 总数 | 活跃数(ratio>0) | 总变体 | 建议 |")
        sb.append("|--------|------------------|----------------|--------|------|")
        for name, total_vfx, active, total_v in p2_candidates:
            sb.append(f"| `{name}` | {total_vfx} | {active} | {total_v:,} | VFX pass 改 `shader_feature_local` / 拆独立 subshader |")
    else:
        sb.append("*未发现需要 VFX 隔离的 shader*")
    sb.append("")

    # ── P3: T3Uber / PostProcessing ──
    sb.append("## P3 — T3Uber / PostProcessing 全 `multi_compile` 优化")
    sb.append("")
    sb.append("> 全部 keyword 使用 `multi_compile` 导致变体组合爆炸")
    sb.append("> 低使用率 keyword 应改为 `shader_feature`")
    sb.append("")

    p3_keywords = ["VIGNETTE", "ENABLE_LIGHT_SHAFT", "LOCAL_EXPOSURE",
                   "ENABLE_FOG", "ENABLE_TONE_MAPPING", "ENABLE_BLOOM"]
    p3_candidates = []
    for s in sorted_shaders:
        # Match T3Uber or PostProcess-like shaders
        name_lower = s.name.lower()
        if not ("uber" in name_lower or "postprocess" in name_lower or "post_process" in name_lower
                or "t3uber" in name_lower):
            continue
        low_usage = []
        for kw in s.keywords:
            locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
            if not locations:
                continue
            avg_r = sum(r for r, _, _ in locations.values()) / len(locations)
            total = sum(t for _, _, t in locations.values())
            mat_usage = material_scan.keyword_material_count.get(kw, 0) if material_scan else -1
            if avg_r < 0.3 or (mat_usage >= 0 and mat_usage <= 3):
                low_usage.append((kw, avg_r, total, mat_usage))
        if low_usage:
            p3_candidates.append((s.name, s.total_variants, len(s.keywords), low_usage))

    if p3_candidates:
        for name, total_v, kw_count, low_usage in p3_candidates:
            sb.append(f"### `{name}` ({kw_count} keywords, {total_v:,} variants)")
            sb.append("")
            sb.append("低使用率 keyword（建议改 `shader_feature`）：")
            sb.append("")
            sb.append("| Keyword | Avg Ratio | 影响变体 | 材质使用数 |")
            sb.append("|---------|-----------|---------|-----------|")
            for kw, avg_r, total, mat in sorted(low_usage, key=lambda x: x[1]):
                mat_str = str(mat) if mat >= 0 else "N/A"
                sb.append(f"| `{kw}` | {avg_r:.3f} | {total:,} | {mat_str} |")
            sb.append("")
    else:
        # Fallback: find any shader with many multi_compile-like keywords (all ratio≈0.5)
        sb.append("*未发现 T3Uber/PostProcessing 相关 shader（或无低使用率 keyword）*")
    sb.append("")

    # ── P4: Stripping rule expansion ──
    sb.append("## P4 — Stripping 规则扩展建议")
    sb.append("")
    sb.append("> 当前 `ShaderPreprocessor.cs` 仅 strip 14 个 shadow/lighting keyword")
    sb.append("> 以下 editor-only / debug keyword 在包体中 ratio = 0，建议新增 strip")
    sb.append("")

    editor_debug_patterns = [
        "SHADER_API_EDITOR", "_DEBUG", "_EDITOR_HLOD_BAKER", "_USE_UNWRAPED_UV",
        "FORCE_DISABLE_SVT", "UNITY_EDITOR", "EDITOR_VISUALIZATION",
        "DEBUG_DISPLAY", "_ENABLE_DEBUG", "SHADER_API_DESKTOP",
    ]
    p4_candidates = []
    for pattern in editor_debug_patterns:
        for s in sorted_shaders:
            for kw in s.keywords:
                if pattern in kw:
                    locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
                    if locations:
                        all_zero = all(r == 0.0 for r, _, _ in locations.values())
                        if all_zero:
                            total = sum(t for _, _, t in locations.values())
                            p4_candidates.append((kw, s.name, total))

    # Deduplicate by keyword
    seen_kw = set()
    p4_unique = []
    for kw, name, total in sorted(p4_candidates, key=lambda x: -x[2]):
        if kw not in seen_kw:
            seen_kw.add(kw)
            # Collect all shaders with this keyword
            affected_shaders = [n for k, n, _ in p4_candidates if k == kw]
            p4_unique.append((kw, affected_shaders, total))

    if p4_unique:
        sb.append("| Keyword | 涉及 Shader 数 | 示例 Shader | 建议 |")
        sb.append("|---------|---------------|------------|------|")
        for kw, shaders_list, total in p4_unique:
            example = ", ".join(f"`{s}`" for s in shaders_list[:3])
            if len(shaders_list) > 3:
                example += f" (+{len(shaders_list) - 3})"
            sb.append(f"| `{kw}` | {len(shaders_list)} | {example} | 加入 ShaderPreprocessor strip 列表 |")
    else:
        sb.append("*未发现可新增 strip 的 editor-only keyword*")
    sb.append("")

    # ── P5: Material keyword dead code ──
    sb.append("## P5 — 材质 Keyword 死代码")
    sb.append("")

    if material_scan and material_scan.all_keywords:
        sb.append("> 材质中启用了但在 shader 中无效的 keyword")
        sb.append("")

        # 5a: Keywords in materials but not in any shader
        orphan_kws = material_scan.all_keywords - all_shader_keywords
        if orphan_kws:
            sb.append("### 孤儿 Keyword（材质启用但不存在于任何 shader）")
            sb.append("")
            sb.append("| Keyword | 材质使用数 |")
            sb.append("|---------|-----------|")
            for kw in sorted(orphan_kws, key=lambda k: -material_scan.keyword_material_count.get(k, 0)):
                count = material_scan.keyword_material_count.get(kw, 0)
                sb.append(f"| `{kw}` | {count} |")
            sb.append("")

        # 5b: Keywords in materials but ratio=0 in the shader (enabled but never compiled ON)
        dead_in_shader = []
        for s in sorted_shaders:
            for kw in s.keywords:
                if kw not in material_scan.all_keywords:
                    continue
                locations = shader_kw_ratios.get(s.name, {}).get(kw, {})
                if locations and all(r == 0.0 for r, _, _ in locations.values()):
                    mat_count = material_scan.keyword_material_count.get(kw, 0)
                    dead_in_shader.append((s.name, kw, mat_count))

        if dead_in_shader:
            sb.append("### 材质启用但 Shader 变体中 ratio=0 的 Keyword")
            sb.append("")
            sb.append("| Shader | Keyword | 材质使用数 | 说明 |")
            sb.append("|--------|---------|-----------|------|")
            for name, kw, mat_count in sorted(dead_in_shader, key=lambda x: -x[2])[:50]:
                sb.append(f"| `{name}` | `{kw}` | {mat_count} | 材质设置无效（keyword 被 strip 或声明为 `shader_feature` 但未编译） |")
            if len(dead_in_shader) > 50:
                sb.append(f"| ... | | | (+{len(dead_in_shader) - 50} more) |")
            sb.append("")
    else:
        sb.append("*未提供 `--project-dir` 参数，跳过材质死代码分析*")
        sb.append("")
        sb.append("> 使用 `--project-dir <unity_project_client_dir>` 启用材质扫描")
        sb.append("")

    with open(os.path.join(output_dir, "optimization_recommendations.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))


def _generate_material_keyword_report(sorted_shaders: List[ShaderInfo], output_dir: str,
                                       material_scan: MaterialScanResult):
    """Generate material_keyword_usage.md — material keyword usage statistics."""
    sb = []
    sb.append("# 材质 Keyword 使用统计")
    sb.append("")
    sb.append(f"**扫描材质数**: {material_scan.total_scanned:,}")
    sb.append(f"**成功解析数**: {material_scan.total_parsed:,}")
    sb.append(f"**不同 Keyword 数**: {len(material_scan.all_keywords):,}")
    sb.append("")

    # All shader keywords for cross-reference
    all_shader_keywords = set()
    for s in sorted_shaders:
        all_shader_keywords.update(s.keywords)

    # Top 50 most used keywords
    sb.append("## Top 50 最常用 Keyword")
    sb.append("")
    sb.append("| # | Keyword | 材质使用数 | Shader 中存在 | 示例材质 |")
    sb.append("|---|---------|-----------|-------------|---------|")
    sorted_kws = sorted(material_scan.keyword_material_count.items(), key=lambda x: -x[1])
    for i, (kw, count) in enumerate(sorted_kws[:50], 1):
        in_shader = "✓" if kw in all_shader_keywords else "✗"
        examples = material_scan.keyword_materials.get(kw, [])[:3]
        ex_str = ", ".join(f"`{os.path.basename(p)}`" for p in examples)
        if len(material_scan.keyword_materials.get(kw, [])) > 3:
            ex_str += " ..."
        sb.append(f"| {i} | `{kw}` | {count:,} | {in_shader} | {ex_str} |")
    sb.append("")

    # Full keyword list
    sb.append("## 全部 Keyword 使用统计")
    sb.append("")
    sb.append("| Keyword | 材质使用数 | Shader 中存在 |")
    sb.append("|---------|-----------|-------------|")
    for kw, count in sorted_kws:
        in_shader = "✓" if kw in all_shader_keywords else "✗"
        sb.append(f"| `{kw}` | {count:,} | {in_shader} |")
    sb.append("")

    # Orphan keywords (in materials but not in any shader keyword list)
    orphan_kws = material_scan.all_keywords - all_shader_keywords
    if orphan_kws:
        sb.append("## 孤儿 Keyword（仅在材质中出现，不存在于任何 Shader）")
        sb.append("")
        sb.append("> 这些 keyword 在材质中被启用，但 shader 编译数据中不包含它们")
        sb.append("> 可能原因：shader 已更新但材质未清理、旧版 keyword 残留")
        sb.append("")
        sb.append("| Keyword | 材质使用数 |")
        sb.append("|---------|-----------|")
        for kw in sorted(orphan_kws, key=lambda k: -material_scan.keyword_material_count.get(k, 0)):
            count = material_scan.keyword_material_count.get(kw, 0)
            sb.append(f"| `{kw}` | {count:,} |")
        sb.append("")

    with open(os.path.join(output_dir, "material_keyword_usage.md"), 'w', encoding='utf-8') as f:
        f.write('\n'.join(sb))


# ════════════════════════════════════════════════════════════════════
# Main Pipeline
# ════════════════════════════════════════════════════════════════════

def scan_jnfs_packages(game_data_dir: str) -> List[Tuple[str, str]]:
    """Find all JNFS idx/data pairs under a game data directory."""
    pairs = []
    streaming_assets = os.path.join(game_data_dir, "StreamingAssets", "res")
    if not os.path.isdir(streaming_assets):
        return pairs

    for root, dirs, files in os.walk(streaming_assets):
        idx_files = [f for f in files if f.endswith('.idx')]
        for idx_file in idx_files:
            idx_path = os.path.join(root, idx_file)
            data_path = idx_path.replace('.idx', '.data')
            if os.path.isfile(data_path):
                pairs.append((idx_path, data_path))

    return sorted(pairs)


def process_jnfs_package(idx_path: str, data_path: str, parser: SerializedFileParser) -> List[ShaderInfo]:
    """Process a single JNFS package and extract shaders."""
    shaders = []
    rel_path = os.path.basename(os.path.dirname(idx_path)) + "/" + os.path.basename(idx_path)
    print(f"  处理 JNFS 包: {rel_path}")

    try:
        data_entries, file_entries = JNFSReader.read_idx(idx_path)
        print(f"    文件数: {len(file_entries)}, 数据块数: {len(data_entries)}")
    except Exception as e:
        print(f"    [ERROR] 读取 idx 失败: {e}", file=sys.stderr)
        return shaders

    # Deduplicate by content hash
    seen_hashes = set()
    hash_to_entry = {e.content_hash: e for e in data_entries}

    for fe in file_entries:
        if fe.content_hash in seen_hashes:
            continue
        seen_hashes.add(fe.content_hash)

        de = hash_to_entry.get(fe.content_hash)
        if de is None:
            continue

        try:
            file_data = JNFSReader.read_file(data_path, de)
        except Exception as e:
            continue

        # Check if this is a SerializedFile (try to parse)
        if len(file_data) < 20:
            continue

        try:
            file_shaders = parser.parse(file_data)
            for s in file_shaders:
                s.source_file = rel_path
                shaders.append(s)
        except Exception:
            pass

    if shaders:
        print(f"    找到 {len(shaders)} 个 Shader")

    return shaders


def process_unity_archive(archive_path: str, parser: SerializedFileParser) -> List[ShaderInfo]:
    """Process a standard UnityFS archive (data.unity3d)."""
    shaders = []
    print(f"  处理 UnityFS: {os.path.basename(archive_path)}")

    try:
        serialized_files = UnityFSReader.read_archive(archive_path)
        print(f"    SerializedFile 数: {len(serialized_files)}")
    except Exception as e:
        print(f"    [ERROR] 读取 UnityFS 失败: {e}", file=sys.stderr)
        return shaders

    for i, sf_data in enumerate(serialized_files):
        try:
            file_shaders = parser.parse(sf_data)
            for s in file_shaders:
                s.source_file = os.path.basename(archive_path)
                shaders.append(s)
        except Exception as e:
            pass

    if shaders:
        print(f"    找到 {len(shaders)} 个 Shader")

    return shaders


def main():
    ap = argparse.ArgumentParser(
        description="JNFS Package Shader Variant Analyzer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data"
  python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" -o ShaderReport
  python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" --jnfs-only
  python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" --unity-only
  python shader_variant_analyzer.py "C:/Users/Admin/Downloads/T3_development/Game_Data" \\
    --project-dir "Project-T3-zhengyang.solis-development/client" -o ShaderReport
        """,
    )
    ap.add_argument("game_data_dir", help="Game data directory (containing data.unity3d and StreamingAssets/)")
    ap.add_argument("-o", "--output", default="ShaderVariantReport", help="Output directory for reports (default: ShaderVariantReport)")
    ap.add_argument("--jnfs-only", action="store_true", help="Only process JNFS packages")
    ap.add_argument("--unity-only", action="store_true", help="Only process data.unity3d")
    ap.add_argument("--dump-jnte", metavar="IDX_PATH", help="Debug: dump JNTE decompression for a specific idx file")
    ap.add_argument("--project-dir", metavar="DIR", help="Unity project client dir for material keyword scanning (enables material analysis)")
    xlsx_group = ap.add_mutually_exclusive_group()
    xlsx_group.add_argument(
        "--xlsx",
        dest="xlsx",
        action="store_true",
        help="导出 Excel (.xlsx) 报告，默认已开启；保留该参数用于兼容旧命令",
    )
    xlsx_group.add_argument(
        "--no-xlsx",
        dest="xlsx",
        action="store_false",
        help="跳过 Excel (.xlsx) 报告导出",
    )
    ap.set_defaults(xlsx=True)
    args = ap.parse_args()

    game_data_dir = args.game_data_dir
    if not os.path.isdir(game_data_dir):
        print(f"Error: {game_data_dir} is not a directory", file=sys.stderr)
        sys.exit(1)

    # Debug mode: dump JNTE
    if args.dump_jnte:
        idx_path = args.dump_jnte
        data_path = idx_path.replace('.idx', '.data')
        print(f"[DEBUG] Dumping JNTE from {idx_path}")
        data_entries, file_entries = JNFSReader.read_idx(idx_path)
        print(f"  DataEntries: {len(data_entries)}, FileEntries: {len(file_entries)}")
        for i, de in enumerate(data_entries[:5]):
            print(f"  DataEntry[{i}]: offset={de.offset}, encoded={de.encoded_size}, decoded={de.decoded_size}")
            try:
                raw = open(data_path, 'rb').read()[de.offset:de.offset + min(de.encoded_size, 64)]
                print(f"    First 64 bytes: {raw.hex()}")
                file_data = JNFSReader.read_file(data_path, de)
                print(f"    Decompressed: {len(file_data)} bytes")
                print(f"    First 32 bytes: {file_data[:32].hex()}")
            except Exception as e:
                print(f"    Error: {e}")
        return

    # Check dependencies
    if zstd is None:
        print("[WARN] zstandard not installed — zstd-compressed JNTE chunks will fail", file=sys.stderr)
        print("  Install: pip install zstandard", file=sys.stderr)
    if lz4 is None:
        print("[WARN] lz4 not installed — LZ4-compressed chunks will fail", file=sys.stderr)
        print("  Install: pip install lz4", file=sys.stderr)

    parser = SerializedFileParser()
    all_shaders: List[ShaderInfo] = []

    print("=" * 60)
    print("Shader Variant Analyzer — JNFS Package Parser")
    print("=" * 60)
    print(f"数据目录: {game_data_dir}")
    print()

    # Process data.unity3d (standard UnityFS)
    if not args.jnfs_only:
        unity3d_path = os.path.join(game_data_dir, "data.unity3d")
        if os.path.isfile(unity3d_path):
            shaders = process_unity_archive(unity3d_path, parser)
            all_shaders.extend(shaders)
        else:
            print(f"  [INFO] data.unity3d 不存在: {unity3d_path}")

    # Process JNFS packages
    if not args.unity_only:
        jnfs_pairs = scan_jnfs_packages(game_data_dir)
        if jnfs_pairs:
            print(f"\n找到 {len(jnfs_pairs)} 个 JNFS 包")
            for idx_path, data_path in jnfs_pairs:
                shaders = process_jnfs_package(idx_path, data_path, parser)
                all_shaders.extend(shaders)
        else:
            print("  [INFO] 未找到 JNFS 包")

    print(f"\n共找到 {len(all_shaders)} 个 Shader")

    # Material keyword scanning
    material_scan = None
    if args.project_dir:
        project_dir = args.project_dir
        if not os.path.isdir(project_dir):
            print(f"[WARN] --project-dir 不存在: {project_dir}", file=sys.stderr)
        else:
            print(f"\n材质 Keyword 扫描: {project_dir}")
            scanner = MaterialKeywordScanner()
            material_scan = scanner.scan_materials(project_dir)

    if all_shaders:
        generate_report(all_shaders, args.output, source_label=game_data_dir,
                       material_scan=material_scan, export_xlsx=args.xlsx)
    else:
        print("未找到 Shader 对象，跳过报告生成")


if __name__ == "__main__":
    main()
